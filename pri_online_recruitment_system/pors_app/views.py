from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.http import JsonResponse, Http404, HttpResponseRedirect, HttpResponse
from django.db.models import Q, Avg, Sum, F
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.generic import TemplateView, DetailView
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db import IntegrityError
from django.core import serializers

#icontains query from a list
#https://thepythonguru.com/python-builtin-functions/reduce/
#https://stackoverflow.com/questions/4824759/django-query-using-contains-each-value-in-a-list
from functools import reduce
#https://docs.python.org/3/library/operator.html
#https://www.geeksforgeeks.org/reduce-in-python/
import operator
from itertools import chain
#for filtering exact salary
from decimal import Decimal

#datetime convertion
import pytz
import datetime
import time
from django.conf import settings



import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color, Protection, colors, GradientFill
from openpyxl.utils import get_column_letter
import csv
import xlwt
from datetime import datetime

from django.forms import modelformset_factory, inlineformset_factory

from .forms import PRIUserProfileRegistrationForm, PRIUserAccountRegistrationForm, PRIClientForm, PRIRequestForm, PRIJobVacancyForm, PRIJobVacancyJobQualificationsForm, PRIJobVacancyJobResponsibilitiesForm, PRIApplicantProfileForm, PRIApplicantSibilingsForm, PRIApplicantEmploymentHistoryForm, PRIApplicantEducationalAttainmentForm, PRIApplicantTrainingsForm, PRIApplicantCharacterReferencesForm, PRIApplicantTest3EssayForm, PRIApplicantTest3SCTForm, PRIApplicantTestSSCTForm, PRIApplicantTestCCATForm, PRIApplicantTestARPForm, PRIClientRequestForm, PRIUserPermissionForm
from .models import PRIUserInfo, PRIUserPermission, PRIClientsInfo, PRIRequestInfo, PRIJobVacancyInfo, PRIJobVacancyJobQualificationsInfo, PRIJobVacancyJobResponsibilitiesInfo, ExaminationInfo, QuestionInfo, ChoicesInfo, PRIApplicantProfileInfo, PRIApplicantSibilingsInfo, PRIApplicantEmploymentHistoryInfo, PRIApplicantEducationalAttainmentInfo, PRIApplicantTrainingsInfo, PRIApplicantCharacterReferencesInfo, PRIApplicantJobRequestInfo, PRIApplicantJobHiredInfo, PRIApplicantTest3EssayInfo,  PRIApplicantTest3SCTInfo, PRIApplicantTestSSCTInfo, PRIApplicantExamScoreT1PEInfo, PRIApplicantExamScoreT2SEInfo, PRIApplicantExamScoreT3EInfo, PRIApplicantExamScoreT3PTSCTInfo, PRIApplicantExamScoreT4ARInfo, PRIApplicantExamScoreCCAInfo, PRIApplicantExamScoreARPInfo, PRIApplicantExamScoreSCCTInfo, PRIApplicantTestCCATInfo, PRIApplicantTestARPInfo


# for encryption
from cryptography.fernet import Fernet
import base64
import logging
import traceback
from django.conf import settings 

# for auto generated applicant id

import random
import string


from django.core.mail import send_mail
# Create your views here.


class PriAdminDashboard(LoginRequiredMixin,TemplateView):
    #dashboard
    template_name = 'pri_admin/pri_admin_index.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriAdminDashboard, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)
          

        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))

            # Policy
            pri_user = PRIUserInfo.objects.all().get(user__in=user)
            policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
            

            no_of_users = PRIUserInfo.objects.all().count()
            no_of_clients = PRIClientsInfo.objects.all().count()
            no_of_hired_applicants = PRIApplicantJobHiredInfo.objects.all().count()
            no_of_request = PRIRequestInfo.objects.all().count()
            no_of_job_vacancy = PRIJobVacancyInfo.objects.all().count()
            no_of_applicants = PRIApplicantProfileInfo.objects.all().count()
            no_of_applying_applicants = PRIApplicantJobRequestInfo.objects.all().count()            
             
        except PRIUserInfo.DoesNotExist:
            raise Http404()

        context.update({
            'user': user,
            'profile': profile,
            'no_of_users': no_of_users,
            'no_of_clients': no_of_clients,
            'no_of_job_vacancy': no_of_job_vacancy,
            'no_of_applicants': no_of_applicants,
            'no_of_request': no_of_request,
            'no_of_applying_applicants': no_of_applying_applicants,
            'no_of_hired_applicants': no_of_hired_applicants,
            'policy': policy,
        })

        return context


"""
=======================================================
*******************For PRI Clients*********************
=======================================================
"""


class PriClients(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_clients_page.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriClients, self).get_context_data(*args, **kwargs)

        user = User.objects.all().filter(username=self.request.user)     

        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))

            # Policy
            pri_user = PRIUserInfo.objects.all().get(user__in=user)
            policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
            if not policy.pri_can_view_clients_page:
                raise Http404()         
        except PRIUserInfo.DoesNotExist:
            raise Http404()

        for key, value in self.request.session.items():
            print('{} -> {}'.format(key, value))

        if 'current_page' in self.request.session:
            del self.request.session['current_page']
            print('current_page => DELETED!')
        else:
            print('current_page is not existing yet!')

        if 'search_text' in self.request.session:
            del self.request.session['search_text']
            print('searched_text => DELETED!')
        else:
            print('searched_text is not existing yet')

        if 'sortBy' in self.request.session:
            del self.request.session['sortBy']
            print('sortBy => DELETED!')
        else:
            print('sortBy is not existing yet')

        if 'column' in self.request.session:
            del self.request.session['column']
            print('column => DELETED')
        else:
            print('column is not existing yet')

        if 'limit' in self.request.session:
            del self.request.session['limit']
            print('limit => DELETED')
        else:
            print('limit is not existing yet')

        query = PRIClientsInfo.objects.all().order_by('-id')
        count = query.count()
        query = pri_admin_client_page_pagination(query, 10, 1)

        columns = []

        for field in PRIClientsInfo._meta.get_fields()[2:]:
            col = field.get_attname_column()[1]
            columns.append(col.replace("_", " "))

        columns.remove("client company logo")

        context.update({
            'user': user,
            'profile': profile,
            'query': query,
            'count': count,
            'columns': columns,
            'policy': policy,
        })

        return context
"""
-------------------------------------------------------
****************PRI Clients Functions******************
-------------------------------------------------------
"""

def pri_admin_client_database_query(searched_text, sortBy, column):
    if sortBy is None or column is None:
        sortable = '-id'
    else:
        if sortBy.lower() == 'Ascending'.lower():
            sortable = column.lower()
        elif sortBy.lower() == 'Descending'.lower():
            sortable = '-' + column.lower()

    if searched_text:
        query = PRIClientsInfo.objects.filter(
            Q(client_user__first_name__icontains=searched_text) |
            Q(client_user__middle_name__icontains=searched_text) |
            Q(client_user__last_name__icontains=searched_text) |
            Q(client_user__user__email__icontains=searched_text) |
            Q(client_location__icontains=searched_text) |
            Q(contact_person__icontains=searched_text) |
            Q(client_company_name__icontains=searched_text) |
            Q(contact_mobile_no__icontains=searched_text)
        ).distinct().order_by(sortable)
    else:
        query = PRIClientsInfo.objects.all().order_by(sortable)

    return query

def pri_admin_client_page_pagination(query, limit, current_page=1):
    if current_page is None:
        current_page = 1

    paginator = Paginator(query, limit)

    try:
        new_query = paginator.page(current_page)
    except PageNotAnInteger:
        print('Error: ', PageNotAnInteger)
        new_query = paginator.page(1)
    except EmptyPage:
        print('Error: ', EmptyPage)
        new_query = paginator.page(paginator.num_pages)

    return new_query

def pri_admin_client_filtered_query(request):
    searched_text = request.session.get('search_text') if request.session.get('search_text') is not None else False
    sortBy = request.session.get('sortBy') if request.session.get('sortBy') is not None else 'Descending'
    column = request.session.get('column') if request.session.get('column') is not None else 'id'
    limit = request.session.get('limit') if request.session.get('limit') is not None else 10
    current_page = request.session.get('current_page') if request.session.get('current_page') is not None else 1

    print('SEARCHED: ', searched_text)
    print('SORT BY: ', sortBy)
    print('COLUMN: ', column.lower().replace(" ", "_"))
    print('LIMIT: ', limit)
    print('CURRENT PAGE: ', current_page)

    query = pri_admin_client_database_query(searched_text, sortBy, column.lower().replace(" ", "_"))
    count = query.count()

    new_query = pri_admin_client_page_pagination(query, limit, current_page)

    return new_query, count

def pri_admin_client_search_filter(request):

    user = User.objects.all().filter(username=request.user)     
    if user:   
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user) 
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))

    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['search_text'] = request.GET.get('search_text')

        new_query, count = pri_admin_client_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': policy,
        }

        data['table_records_client'] = render_to_string('pri_admin/tables/pri_admin_clients_table.html', context)
        data['pagination_client'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

def pri_admin_client_record_limiter(request):

    user = User.objects.all().filter(username=request.user)     
    if user:   
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user) 
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))

    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['limit'] = request.GET.get('limit')

        new_query, count = pri_admin_client_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': policy,
        }

        data['table_records_client'] = render_to_string('pri_admin/tables/pri_admin_clients_table.html', context)
        data['pagination_client'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

def pri_admin_client_sort_records(request):
    user = User.objects.all().filter(username=request.user)     
    if user:   
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)      
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))
    if request.method == 'GET' and request.is_ajax():
        data = dict()

        request.session['sortBy'] = request.GET.get('sortBy')
        request.session['column'] = request.GET.get('column')

        new_query, count = pri_admin_client_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy':policy,
        }

        data['table_records_client'] = render_to_string('pri_admin/tables/pri_admin_clients_table.html', context)
        data['pagination_client'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

def pri_admin_client_paging(request):
    user = User.objects.all().filter(username=request.user)     
    if user:   
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user) 
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['current_page'] = request.GET.get('page')

        new_query, count = pri_admin_client_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': policy,
        }

        data['table_records_client'] = render_to_string('pri_admin/tables/pri_admin_clients_table.html', context)
        data['pagination_client'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()


"""
-------------------------------------------------------
*****************PRI Clients Create********************
-------------------------------------------------------
"""

def pri_admin_create_clients(request):
    data = dict()
    template_name = 'pri_admin/pri_admin_clients/pri_admin_create_clients.html'

    user = User.objects.all().filter(username=request.user)      

    if user:  
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_add_clients:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))

    if request.is_ajax():
        if request.method == 'POST':
            form = PRIClientForm(request.POST, request.FILES or None)
            if form.is_valid():

                instance = form.save(commit=False)

                if 'client_company_logo' in request.FILES:
                    instance.client_company_logo = request.FILES['client_company_logo']
                else:
                    print("NO IMAGE")

                instance.save()

                new_query, count = pri_admin_client_filtered_query(request)

                context = {
                    'query': new_query,
                    'count': count,
                    'policy': policy,
                }

                data['table_records_client'] = render_to_string('pri_admin/tables/pri_admin_clients_table.html', context)
                data['pagination_client'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        elif request.method == 'GET':
            form = PRIClientForm(request.FILES or None)

        context = {
            'form': form,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()



def pri_admin_edit_clients(request, id):
    data = dict()
    client = get_object_or_404(PRIClientsInfo, id=id)
    template_name = 'pri_admin/pri_admin_clients/pri_admin_edit_clients.html'

    user = User.objects.all().filter(username=request.user)
    if user:        
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_edit_clients:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))

    if request.is_ajax():
        if request.method == 'POST':
            form = PRIClientForm(request.POST, request.FILES or None, instance=client )
            if form.is_valid():

                instance = form.save(commit=False)

                if 'client_company_logo' in request.FILES:
                    instance.client_company_logo = request.FILES['client_company_logo']
                else:
                    print("NO IMAGE")

                instance.save()

                new_query, count = pri_admin_client_filtered_query(request)

                context = {
                    'query': new_query,
                    'count': count,
                    'policy': policy,
                }

                data['table_records_client'] = render_to_string('pri_admin/tables/pri_admin_clients_table.html', context)
                data['pagination_client'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        elif request.method == 'GET':
            form = PRIClientForm(request.FILES or None, instance=client)

        context = {
            'form': form,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()

def pri_admin_delete_clients(request, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_clients/pri_admin_delete_clients.html'
    client = get_object_or_404(PRIClientsInfo, id=id)

    user = User.objects.all().filter(username=request.user)       
    if user: 
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_delete_clients:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))


    if request.is_ajax():
        if request.method == "POST":

            client.delete()

            new_query, count = pri_admin_client_filtered_query(request)

            context = {
                'query': new_query,
                'count': count,
                'policy': policy,
            }

            data['table_records_client'] = render_to_string('pri_admin/tables/pri_admin_clients_table.html', context)
            data['pagination_client'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)
            data['form_is_valid'] = True

        elif request.method == "GET":
            context = {
                'client': client,
            }

            data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()
# Show client applicants

def pri_admin_client_show_applicants(request, id): 
    template_name = 'pri_admin/pri_admin_clients/pri_admin_show_client_applicants.html'
    client = get_object_or_404(PRIClientsInfo, id=id)
    applicants = PRIApplicantJobHiredInfo.objects.all().filter(company_name=client).order_by('-id').distinct()

    user = User.objects.all().filter(username=request.user) 

    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))

        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_hired_clients:
            raise Http404()      
    except PRIUserInfo.DoesNotExist:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))
        # raise Http404()
 
    context = {
        'user': user,
        'profile': profile,
        'client': client,
        'applicants': applicants,
        'policy': policy,
    }

    return render(request, template_name, context)

def pri_admin_client_hired_applicant_set_status(request, id):
    data = dict()

    hired_applicant = get_object_or_404(PRIApplicantJobHiredInfo, id=id)

    if request.is_ajax():
        if request.method == "GET":
            pass
        elif request.method == "POST":
            status = json.loads(request.body)
            # print in readable format
            print(json.dumps(status, indent=4, sort_keys=True))
            hired_applicant.status = status['status']
            hired_applicant.save()
            data['confirmation'] = "Record Updated Successfully"
        return JsonResponse(data)
    else:
        raise Http404("Page Not Found")

def pri_admin_client_delete_applicants(request, id, aid):
    data = dict()
    template_name = 'pri_admin/pri_admin_clients/pri_admin_delete_client_applicants.html'
    client = get_object_or_404(PRIClientsInfo, id=id)
    applicants = PRIApplicantJobHiredInfo.objects.all().filter(company_name=client).order_by('-id').distinct()
    
    hired_applicant = get_object_or_404(PRIApplicantJobHiredInfo, id=aid)

    user = User.objects.all().filter(username=request.user)

    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))
    except PRIUserInfo.DoesNotExist:
        raise Http404()
        
    
    if request.is_ajax():
        if request.method == 'POST': 
            hired_applicant.delete()
            context = {
                'user': user,
                'profile': profile,
                'client': client,
                'hired_applicant': hired_applicant,
                'applicants': applicants, 
            }
            data['table_records'] = render_to_string('pri_admin/tables/pri_admin_show_client_applicants_table.html', context)

            data['form_is_valid'] = True

        elif request.method == 'GET':
            context = {
                'user': user,
                'profile': profile,
                'client': client,
                'hired_applicant': hired_applicant,
                'applicants': applicants, 
            }
            data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

"""
=====================================================
*******************END PRI CLIENTS*******************
=====================================================
"""

"""
For PRI Hired
"""


class PriHired(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_hired_page.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriHired, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)

        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))
        except PRIUserInfo.DoesNotExist:
            raise Http404()

        context.update({
            'user': user,
            'profile': profile
        })

        return context


"""
=======================================================================
***************************For PRI Jobs******************************** 
=======================================================================
"""
# For I encryption
# Follow this steps
# https://www.pythoncircle.com/post/641/encryption-decryption-in-python-django/
def encrypt_key(txt):
    try:
        # convert integer etc to string first   
        txt = str(txt)
        # get key from settings
        cipher_suite = Fernet(settings.ENCRYPT_KEY) # key should be byte
        # #input should be byte, so convert the text to byte
        encrypted_text = cipher_suite.encrypt(txt.encode('ascii'))
        # encode to urlsafe base64 format
        encrypted_text =base64.urlsafe_b64encode(encrypted_text).decode("ascii")

        return encrypted_text

    except Exception as e:
        # log if an error
        logging.getLogger("error_logger").error(traceback.format_exc())
        return None

def decrypt_key(txt):
    try:
        # base64 decode
        txt = base64.urlsafe_b64decode(txt)
        cipher_suite = Fernet(settings.ENCRYPT_KEY)
        decoded_text = cipher_suite.decrypt(txt).decode("ascii")
        return decoded_text
    except Exception as e:
    # log the error
        logging.getLogger("error_logger").error(traceback.format_exc())
        return None

# print(encrypt_key(1))
# print(decrypt_key(encrypt_key(1)))
 

# For Viewing applicants from every job posts request
class PriJobsApplicantsRequests(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_jobs_applicants_page.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriJobsApplicantsRequests, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)

        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))
            #job_vacancies = get_object_or_404(PRIJobVacancyInfo, secure_id=self.kwargs.get('key'))
            # or
            key = decrypt_key(self.kwargs.get('key'))
            job_vacancies = get_object_or_404(PRIJobVacancyInfo, id=key)            
            job_requests = PRIApplicantJobRequestInfo.objects.all().filter(job_vacancy_applied=job_vacancies)
            # # PRIApplicantJobRequestInfo
            # print(job_vacancies.secure_id)
            # print(decrypt_key(job_vacancies.secure_id))
            # print(self.kwargs.get('key'))
            # print(decrypt_key(self.kwargs.get('key')))            

            # Policy
            pri_user = PRIUserInfo.objects.all().get(user__in=user)
            policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
            if not policy.pri_can_view_job_vacancy_applicants:
                raise Http404()         

        except PRIUserInfo.DoesNotExist:
            raise Http404()

        context.update({
            'user': user,
            'profile': profile, 
            'job_vacancies': job_vacancies,
            'job_requests': job_requests,
            'policy':policy,
        })

        return context 
# Auto generating applicant id
def randomApplicantId(string_length=7):
# https://stackoverflow.com/questions/2511222/efficiently-generate-a-16-character-alphanumeric-string
# https://pynative.com/python-generate-random-string/
    numbers = string.digits
    letters = string.ascii_uppercase

    id = ''.join(random.choice(letters + numbers) for i in range(string_length))

    return id

@login_required
def pri_applicant_profile_view(request, id, akey):
    template_name = 'pri_admin/pri_admin_applicant_profile_viewer.html'

    user = User.objects.all().filter(username=request.user)

    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))       
                           
    except PRIUserInfo.DoesNotExist:
        raise Http404() 

    key = decrypt_key(akey)  
    applicants = PRIApplicantProfileInfo.objects.all().filter(id=key)
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=id)  
    job_request = PRIApplicantJobRequestInfo.objects.all().filter(job_vacancy_applied=job_vacancy, applying_applicant__in=applicants)
    
    try:
        applicant_job_request = PRIApplicantJobRequestInfo.objects.all().get(job_vacancy_applied=job_vacancy, applying_applicant__in=applicants)
    except PRIApplicantJobRequestInfo.DoesNotExist:
        raise Http404("Applicant Details Not Found!")
     
    t1pe_score = PRIApplicantExamScoreT1PEInfo.objects.all().filter(applicant_exam_score_t1pe__in=applicants, job_request_exam_score_t1pe__in=job_request)
    t2se_score = PRIApplicantExamScoreT2SEInfo.objects.all().filter(applicant_exam_score_t2se__in=applicants, job_request_exam_score_t2se__in=job_request)
    t3e_score = PRIApplicantExamScoreT3EInfo.objects.all().filter(applicant_exam_score_t3e__in=applicants, job_request_exam_score_t3e__in=job_request)
    t3ptsct_score = PRIApplicantExamScoreT3PTSCTInfo.objects.all().filter(applicant_exam_score_t3ptsct__in=applicants, job_request_exam_score_t3ptsct__in=job_request)
    t4ar_score = PRIApplicantExamScoreT4ARInfo.objects.all().filter(applicant_exam_score_t4ar__in=applicants, job_request_exam_score_t4ar__in=job_request)
    cca_score = PRIApplicantExamScoreCCAInfo.objects.all().filter(applicant_exam_score_cca__in=applicants, job_request_exam_score_cca__in=job_request)
    arp_score = PRIApplicantExamScoreARPInfo.objects.all().filter(applicant_exam_score_arp__in=applicants, job_request_exam_score_arp__in=job_request)
    scct_score = PRIApplicantExamScoreSCCTInfo.objects.all().filter(applicant_exam_score_scct__in=applicants, job_request_exam_score_scct__in=job_request)
    
    final_score = 0
    final_over = 0

    for t1pe in t1pe_score: 
        final_score += t1pe.score_t1pe
        final_over += t1pe.over_t1pe

    for t2se in t2se_score: 
        final_score += t2se.score_t2se
        final_over += t2se.over_t2se
    
    for t3e in t3e_score: 
        final_score += t3e.score_t3e
        final_over += t3e.over_t3e
    
    for t3ptsct in t3ptsct_score: 
        final_score += t3ptsct.score_t3ptsct
        final_over += t3ptsct.over_t3ptsct
    
    for t4ar in t4ar_score: 
        final_score += t4ar.score_t4ar
        final_over += t4ar.over_t4ar
    
    for cca in cca_score: 
        final_score += cca.score_cca
        final_over += cca.over_cca
    
    for arp in arp_score: 
        final_score += arp.score_arp
        final_over += arp.over_arp
    
    for scct in scct_score: 
        final_score += scct.score_scct
        final_over += scct.over_scct
     
    pri_request = PRIRequestInfo.objects.all().get(job_vacancy_client_request_fk=job_vacancy) 
    pri_client = PRIClientsInfo.objects.all().get(requested_client_fk=pri_request)
 
    if request.method == 'POST':
        
        for applicant in applicants:
            hired, created = PRIApplicantJobHiredInfo.objects.update_or_create(company_name=pri_client, hired_applicant=applicant, 
                defaults={'applicant_id': randomApplicantId(7)})

            applicant.position_desired = job_vacancy.job_title
            applicant.company_assigned = pri_client.client_company_name
            applicant.save()
            print('Company Assigned', pri_client.client_company_name)
            print('Position',job_vacancy.job_title)
            # Send email
            applicant_email = User.objects.all().get(applicant_user_fk=applicant)
            from_email = settings.DEFAULT_FROM_EMAIL
            subject = 'PRI Online Hiring System E-mail Notification'
            body_of_the_message = "Congratulations you've been hired by ({}) and will be assigned on ({}) with the applied postion ({})".format("PRI", pri_client.client_company_name, job_vacancy.job_title)
            recipient_email = applicant_email.email
            send_mail(subject, body_of_the_message, from_email, [recipient_email], fail_silently=True) # this is worrking

        return HttpResponseRedirect(reverse_lazy('pri:applicant_request_jobs', kwargs={
            'key': job_vacancy.secure_id
        }))

    context = {
        'applicants': applicants,
        't1pe_score': t1pe_score,
        't2se_score': t2se_score,
        't3e_score': t3e_score,
        't3ptsct_score': t3ptsct_score,
        't4ar_score': t4ar_score,
        'cca_score': cca_score,
        'arp_score': arp_score,
        'scct_score': scct_score,
        'final_score': final_score,
        'final_over': final_over,
        'final_passing_score': (final_over/2),
        'job_request': job_request,
        'job_vacancy': job_vacancy,
    }
    
    return render(request, template_name, context)
# Reviewing applicant examninations
@login_required
def pri_applicant_test3_essay_review(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exam_review/pri_admin_applicant_exam_test3_review.html'
    user = User.objects.all().filter(username=request.user)
    
    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))       
                           
    except PRIUserInfo.DoesNotExist:
        raise Http404() 

    key = decrypt_key(key)  
    applicant = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=id)  
    job_request = PRIApplicantJobRequestInfo.objects.all().get(job_vacancy_applied=job_vacancy, applying_applicant=applicant)
     
    #essayTest = PRIApplicantTest3EssayInfo.objects.all().exclude('id').filter(applicant_exam=applicant, job_request_exam=job_request).values_list().distinct()
    
    # list_question = []
    # list_answers = []
    # fields = []
    # for field in PRIApplicantTest3EssayInfo._meta.get_fields()[3:]:
    #     field_name = field.get_attname_column()[1] 
    #     fields.append(field_name)
    #     verbose_name = PRIApplicantTest3EssayInfo._meta.get_field(field_name).verbose_name    
    #     record = PRIApplicantTest3EssayInfo.objects.all().filter(applicant_exam=applicant, job_request_exam=job_request).values_list(field_name,flat=True).distinct()
    #     for rec in record:
    #         list_question.append(field_name,rec)
    

    # print(list_question)

    if request.method == 'GET':
        essayTest = PRIApplicantTest3EssayInfo.objects.all().get(applicant_exam=applicant, job_request_exam=job_request)        
        essayForm = PRIApplicantTest3EssayForm(request.GET or None, instance=essayTest)        
        context = {
            'applicant': applicant, 
            'essayForm': essayForm,
        }

        return render(request, template_name, context)

    elif request.method == 'POST':
        score = request.POST['score']
        enable_retake = request.POST.getlist('enable-retake')
        allow_retake = False
        if enable_retake:
            value = enable_retake[0]
            if value == 'on':
                allow_retake = True
        
        print('----------->', score)
        over = 30
        ratings, status = calculate_exam_score(int(score), over)

        t3e, created = PRIApplicantExamScoreT3EInfo.objects.update_or_create(applicant_exam_score_t3e=applicant, job_request_exam_score_t3e=job_request, 
            defaults={'score_t3e': int(score), 'over_t3e' : over, 'status_t3e': status, 'ratings_t3e': ratings, 'allow_retake_t3e': allow_retake})


        return HttpResponseRedirect(reverse_lazy('pri:applicant_profile_viewer', kwargs={
                'id': job_vacancy.id,
                'akey': applicant.secure_key_id
            }))


@login_required
def pri_applicant_test3_ptsct_review(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exam_review/pri_admin_applicant_exam_test3_ptsct_review.html'
    user = User.objects.all().filter(username=request.user)
    
    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))       
                           
    except PRIUserInfo.DoesNotExist:
        raise Http404() 

    key = decrypt_key(key)  
    applicant = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=id)  
    job_request = PRIApplicantJobRequestInfo.objects.all().get(job_vacancy_applied=job_vacancy, applying_applicant=applicant)
      

    if request.method == 'GET':
        t3ptsctTest = PRIApplicantTest3SCTInfo.objects.all().get(applicant_exam_sct=applicant, job_request_exam_sct=job_request)        
        t3ptsctForm = PRIApplicantTest3SCTForm(request.GET or None, instance=t3ptsctTest)        
        context = {
            'applicant': applicant, 
            't3ptsctForm': t3ptsctForm,
        }

        return render(request, template_name, context)

    elif request.method == 'POST':
        score = request.POST['score']
        enable_retake = request.POST.getlist('enable-retake')
        allow_retake = False
        if enable_retake:
            value = enable_retake[0]
            if value == 'on':
                allow_retake = True
        
        print('----------->', score)
        over = 15
        ratings, status = calculate_exam_score(int(score), over)

        t3ptsct, created = PRIApplicantExamScoreT3PTSCTInfo.objects.update_or_create(applicant_exam_score_t3ptsct=applicant, job_request_exam_score_t3ptsct=job_request, 
            defaults={'score_t3ptsct': int(score), 'over_t3ptsct' : over, 'status_t3ptsct': status, 'ratings_t3ptsct': ratings, 'allow_retake_t3ptsct': allow_retake})


        return HttpResponseRedirect(reverse_lazy('pri:applicant_profile_viewer', kwargs={
                'id': job_vacancy.id,
                'akey': applicant.secure_key_id
            })) 

@login_required
def pri_applicant_test_ccat_review(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exam_review/pri_admin_applicant_exam_test_ccat_review.html'
    user = User.objects.all().filter(username=request.user)
    
    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))       
                           
    except PRIUserInfo.DoesNotExist:
        raise Http404() 

    key = decrypt_key(key)  
    applicant = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=id)  
    job_request = PRIApplicantJobRequestInfo.objects.all().get(job_vacancy_applied=job_vacancy, applying_applicant=applicant)
      

    if request.method == 'GET':
        ccatTest = PRIApplicantTestCCATInfo.objects.all().get(applicant_exam_ccat=applicant, job_request_exam_ccat=job_request)        
        ccatForm = PRIApplicantTestCCATForm(request.GET or None, instance=ccatTest)        
        context = {
            'applicant': applicant, 
            'ccatForm': ccatForm,
        }

        return render(request, template_name, context)

    elif request.method == 'POST':
        score = request.POST['score']
        enable_retake = request.POST.getlist('enable-retake')
        allow_retake = False
        if enable_retake:
            value = enable_retake[0]
            if value == 'on':
                allow_retake = True
        
        print('----------->', score)
        over = 60
        ratings, status = calculate_exam_score(int(score), over)

        ccat, created = PRIApplicantExamScoreCCAInfo.objects.update_or_create(applicant_exam_score_cca=applicant, job_request_exam_score_cca=job_request, 
            defaults={'score_cca': int(score), 'over_cca' : over, 'status_cca': status, 'ratings_cca': ratings, 'allow_retake_cca': allow_retake})


        return HttpResponseRedirect(reverse_lazy('pri:applicant_profile_viewer', kwargs={
                'id': job_vacancy.id,
                'akey': applicant.secure_key_id
            }))


@login_required
def pri_applicant_test_arp_review(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exam_review/pri_admin_applicant_exam_test_arp_review.html'
    user = User.objects.all().filter(username=request.user)
    
    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))       
                           
    except PRIUserInfo.DoesNotExist:
        raise Http404() 

    key = decrypt_key(key)  
    applicant = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=id)  
    job_request = PRIApplicantJobRequestInfo.objects.all().get(job_vacancy_applied=job_vacancy, applying_applicant=applicant)
      

    if request.method == 'GET':
        arpTest = PRIApplicantTestARPInfo.objects.all().get(applicant_exam_arp=applicant, job_request_exam_arp=job_request)        
        arpForm = PRIApplicantTestARPForm(request.GET or None, instance=arpTest)        
        context = {
            'applicant': applicant, 
            'arpForm': arpForm,
        }

        return render(request, template_name, context)

    elif request.method == 'POST':
        score = request.POST['score']
        enable_retake = request.POST.getlist('enable-retake')
        allow_retake = False
        if enable_retake:
            value = enable_retake[0]
            if value == 'on':
                allow_retake = True
        
        print('----------->', score)
        over = 65
        ratings, status = calculate_exam_score(int(score), over)

        ccat, created = PRIApplicantExamScoreARPInfo.objects.update_or_create(applicant_exam_score_arp=applicant, job_request_exam_score_arp=job_request, 
            defaults={'score_arp': int(score), 'over_arp' : over, 'status_arp': status, 'ratings_arp': ratings, 'allow_retake_arp': allow_retake})


        return HttpResponseRedirect(reverse_lazy('pri:applicant_profile_viewer', kwargs={
                'id': job_vacancy.id,
                'akey': applicant.secure_key_id
            })) 

@login_required
def pri_applicant_test_scct_review(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exam_review/pri_admin_applicant_exam_test_ssct_review.html'
    user = User.objects.all().filter(username=request.user)
    
    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))       
                           
    except PRIUserInfo.DoesNotExist:
        raise Http404() 

    key = decrypt_key(key)  
    applicant = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=id)  
    job_request = PRIApplicantJobRequestInfo.objects.all().get(job_vacancy_applied=job_vacancy, applying_applicant=applicant)
      

    if request.method == 'GET':
        scctTest = PRIApplicantTestSSCTInfo.objects.all().get(applicant_exam_ssct=applicant, job_request_exam_ssct=job_request)        
        scctForm = PRIApplicantTestSSCTForm(request.GET or None, instance=scctTest)        
        context = {
            'applicant': applicant, 
            'scctForm': scctForm,
        }

        return render(request, template_name, context)

    elif request.method == 'POST':
        score = request.POST['score']
        enable_retake = request.POST.getlist('enable-retake')
        allow_retake = False
        if enable_retake:
            value = enable_retake[0]
            if value == 'on':
                allow_retake = True
        
        print('----------->', score)
        over = 60
        ratings, status = calculate_exam_score(int(score), over)

        ccat, created = PRIApplicantExamScoreSCCTInfo.objects.update_or_create(applicant_exam_score_scct=applicant, job_request_exam_score_scct=job_request, 
            defaults={'score_scct': int(score), 'over_scct' : over, 'status_scct': status, 'ratings_scct': ratings, 'allow_retake_scct': allow_retake})


        return HttpResponseRedirect(reverse_lazy('pri:applicant_profile_viewer', kwargs={
                'id': job_vacancy.id,
                'akey': applicant.secure_key_id
            })) 

@login_required
def pri_applicant_test1_review(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exam_review/pri_admin_applicant_exam_test1_review.html'
    user = User.objects.all().filter(username=request.user)
    
    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))       
                           
    except PRIUserInfo.DoesNotExist:
        raise Http404() 

    key = decrypt_key(key)  
    applicant = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=id)  
    job_request = PRIApplicantJobRequestInfo.objects.all().get(job_vacancy_applied=job_vacancy, applying_applicant=applicant)
      

    if request.method == 'GET':
        test1Test = PRIApplicantExamScoreT1PEInfo.objects.all().get(applicant_exam_score_t1pe=applicant, job_request_exam_score_t1pe=job_request)        
        
        
        context = {
            'applicant': applicant,
            'test1Test': test1Test,
        }

        return render(request, template_name, context)

    elif request.method == 'POST': 
        is_retakable = request.POST.getlist('is-retake')
        allow_retake = False
        if is_retakable:
            value = is_retakable[0]
            if value == 'on':
                allow_retake = True 
        print('------->',is_retakable)
        t1, created = PRIApplicantExamScoreT1PEInfo.objects.update_or_create(applicant_exam_score_t1pe=applicant, job_request_exam_score_t1pe=job_request, 
            defaults={'allow_retake_t1pe': allow_retake})


        return HttpResponseRedirect(reverse_lazy('pri:applicant_profile_viewer', kwargs={
                'id': job_vacancy.id,
                'akey': applicant.secure_key_id
            })) 

@login_required
def pri_applicant_test2_review(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exam_review/pri_admin_applicant_exam_test2_review.html'
    user = User.objects.all().filter(username=request.user)
    
    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))       
                           
    except PRIUserInfo.DoesNotExist:
        raise Http404() 

    key = decrypt_key(key)  
    applicant = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=id)  
    job_request = PRIApplicantJobRequestInfo.objects.all().get(job_vacancy_applied=job_vacancy, applying_applicant=applicant)
      

    if request.method == 'GET':
        test2Test = PRIApplicantExamScoreT2SEInfo.objects.all().get(applicant_exam_score_t2se=applicant, job_request_exam_score_t2se=job_request)        
        
        
        context = {
            'applicant': applicant,
            'test2Test': test2Test,
        }

        return render(request, template_name, context)

    elif request.method == 'POST': 
        is_retakable = request.POST.getlist('is-retake')
        allow_retake = False
        if is_retakable:
            value = is_retakable[0]
            if value == 'on':
                allow_retake = True 
        print('------->',is_retakable)
        t1, created = PRIApplicantExamScoreT2SEInfo.objects.update_or_create(applicant_exam_score_t2se=applicant, job_request_exam_score_t2se=job_request, 
            defaults={'allow_retake_t2se': allow_retake})


        return HttpResponseRedirect(reverse_lazy('pri:applicant_profile_viewer', kwargs={
                'id': job_vacancy.id,
                'akey': applicant.secure_key_id
            })) 

@login_required
def pri_applicant_test4_review(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exam_review/pri_admin_applicant_exam_test4_review.html'
    user = User.objects.all().filter(username=request.user)
    
    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))       
                           
    except PRIUserInfo.DoesNotExist:
        raise Http404() 

    key = decrypt_key(key)  
    applicant = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=id)  
    job_request = PRIApplicantJobRequestInfo.objects.all().get(job_vacancy_applied=job_vacancy, applying_applicant=applicant)
      

    if request.method == 'GET':
        test4Test = PRIApplicantExamScoreT4ARInfo.objects.all().get(applicant_exam_score_t4ar=applicant, job_request_exam_score_t4ar=job_request)        
        
        
        context = {
            'applicant': applicant,
            'test4Test': test4Test,
        }

        return render(request, template_name, context)

    elif request.method == 'POST': 
        is_retakable = request.POST.getlist('is-retake')
        allow_retake = False
        if is_retakable:
            value = is_retakable[0]
            if value == 'on':
                allow_retake = True 
        print('------->',is_retakable)
        t4, created = PRIApplicantExamScoreT4ARInfo.objects.update_or_create(applicant_exam_score_t4ar=applicant, job_request_exam_score_t4ar=job_request, 
            defaults={'allow_retake_t4ar': allow_retake})


        return HttpResponseRedirect(reverse_lazy('pri:applicant_profile_viewer', kwargs={
                'id': job_vacancy.id,
                'akey': applicant.secure_key_id
            })) 

# Enable exams for applicants
# Ajax request receiver
def pri_jobs_applicant_requests_toggle_exam(request, key, id):
    data = dict()

    job_request = get_object_or_404(PRIApplicantJobRequestInfo, id=id)
    job_vacancies = get_object_or_404(PRIJobVacancyInfo, id=decrypt_key(key))
    job_requests = PRIApplicantJobRequestInfo.objects.all().filter(job_vacancy_applied=job_vacancies)

    if request.is_ajax():
        if request.method == 'GET':        
            print("Hello")
        elif request.method == 'POST':
            """
            HttpRequest.body
            The raw HTTP request body as a bytestring. This is useful for processing data in 
            different ways than conventional HTML forms: binary images, XML payload etc. For 
            processing conventional form data, use HttpRequest.POST.

            """
            #print(request.body)
            # load the json object
            datas = json.loads(request.body) 
            # print in readable format
            print(json.dumps(datas, indent=4, sort_keys=True))
            # accessing through every object 
            is_exam = datas['is_exam'] 
            if is_exam:
                job_request.take_exam = False
                job_request.save()
            else:
                job_request.take_exam = True
                job_request.save()  
            
            context = {
                'job_vacancies': job_vacancies,
                'job_requests': job_requests,
            }

            data['table_records_job_applicant_requests'] = render_to_string('pri_admin/tables/pri_admin_jobs_applicant_request_table.html', context)
            
        return JsonResponse(data, safe=False)
    else:
        raise Http404()
# Set Interview date schedule
def pri_jobs_applicant_requests_set_interview_sched(request, key, id):
    data = dict()
    job_request = get_object_or_404(PRIApplicantJobRequestInfo, id=id)
    job_vacancies = get_object_or_404(PRIJobVacancyInfo, id=decrypt_key(key))
    job_requests = PRIApplicantJobRequestInfo.objects.all().filter(job_vacancy_applied=job_vacancies)


    if request.is_ajax():
        if request.method == 'POST':
            datas = json.loads(request.body)

            print(json.dumps(datas, indent=4, sort_keys=True))
            date = datas['date']
            #YYYY-MM-DD HH:MM[:ss[.uuuuuu]][TZ] format."] 
          
            old_format = datetime.strptime(date, '%b %d %Y %I:%M %p')
            tz_aware_datetetime = old_format.replace(tzinfo=pytz.timezone(settings.TIME_ZONE))
            standard_format = tz_aware_datetetime.strftime('%Y-%m-%d %H:%M:%S') 
            

            
            print(standard_format)
            
            job_request.interview_date = standard_format
            job_request.save()
        
        context = {
            'job_vacancies': job_vacancies,
            'job_requests': job_requests,
        }
        
        data['table_records_job_applicant_requests'] = render_to_string('pri_admin/tables/pri_admin_jobs_applicant_request_table.html', context)
        return JsonResponse(data,  safe=False)        
    else:
        raise Http404()
    

class PriJobs(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_jobs_page.html'

    """
    
    # For searching
    # https://targetjobs.co.uk/careers-advice/job-descriptions
    # https://stackoverflow.com/questions/4843158/check-if-a-python-list-item-contains-a-string-inside-another-string
    list_of_data = ["asdasdas","qweqwe","zxczxc"]

    for s in list_of_data: 
        if "zxc" in s:
            print("YEs")
            print(s)
        else:
            print("No")

    if any(["a" in match for match in list_of_data]):
        print(True)
        
    """

    def get_context_data(self, *args, **kwargs):
        context = super(PriJobs, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)

        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))

            # Policy
            pri_user = PRIUserInfo.objects.all().get(user__in=user)
            policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
            if not policy.pri_can_view_job_vacancy_page:
                raise Http404()      
        except PRIUserInfo.DoesNotExist:
            raise Http404()

        for key, value in self.request.session.items():
            print('{} -> {}'.format(key, value))

        if 'current_page' in self.request.session:
            del self.request.session['current_page']
            print('current_page => DELETED!')
        else:
            print('current_page is not existing yet!')

        if 'search_text' in self.request.session:
            del self.request.session['search_text']
            print('searched_text => DELETED!')
        else:
            print('searched_text is not existing yet')

        if 'sortBy' in self.request.session:
            del self.request.session['sortBy']
            print('sortBy => DELETED!')
        else:
            print('sortBy is not existing yet')

        if 'column' in self.request.session:
            del self.request.session['column']
            print('column => DELETED')
        else:
            print('column is not existing yet')

        if 'limit' in self.request.session:
            del self.request.session['limit']
            print('limit => DELETED')
        else:
            print('limit is not existing yet')

        query = PRIJobVacancyInfo.objects.all().order_by('-id')
        count = query.count()
        query = pri_admin_jobs_page_pagination(query, 10, 1)

        columns = []

        for field in PRIJobVacancyInfo._meta.get_fields()[3:]:
            col = field.get_attname_column()[1]
            columns.append(col.replace("_", " "))

        columns.remove("client request id")
        columns.remove("job company overview")
        columns.remove("job description")

        # testQ = PRIJobVacancyInfo.objects.annotate(fieldsum=F(1) + F(2)).order_by('fieldsum')

        test = PRIJobVacancyInfo.objects.extra(  select={'fieldsum':2 + 2},
        order_by=('fieldsum',))
        
        # job_history = PRIApplicantEmploymentHistoryInfo.objects.filter

        print(test)

        context.update({
            'user': user,
            'profile': profile,
            'query': query,
            'count': count,
            'columns': columns,
            'policy': policy,
        })

        return context

def pri_admin_create_jobs(request, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_jobs/pri_admin_create_jobs.html'
    requests_post = get_object_or_404(PRIRequestInfo, id=id)

    user = User.objects.all().filter(username=request.user)        
    if user:
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_create_job_request:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))


    #For edit
    #PRIJobVacancyJobQualificationsFormset = modelformset_factory(PRIJobVacancyJobQualificationsInfo, form=PRIJobVacancyJobQualificationsForm, extra=1, can_delete=True)
    #PRIJobVacancyJobResponsibilitiesFormset = modelformset_factory(PRIJobVacancyJobResponsibilitiesInfo, form=PRIJobVacancyJobResponsibilitiesForm, extra=1,  can_delete=True)

    PRIJobVacancyJobQualificationsFormset = modelformset_factory(PRIJobVacancyJobQualificationsInfo, form=PRIJobVacancyJobQualificationsForm, extra=1,)
    PRIJobVacancyJobResponsibilitiesFormset = modelformset_factory(PRIJobVacancyJobResponsibilitiesInfo, form=PRIJobVacancyJobResponsibilitiesForm, extra=1, )

    if request.is_ajax():
        if request.method == 'POST':
            form1 = PRIJobVacancyForm(request.POST or None)
            form2 = PRIJobVacancyJobQualificationsFormset(request.POST or None,  queryset=PRIJobVacancyJobQualificationsInfo.objects.none())
            form3 = PRIJobVacancyJobResponsibilitiesFormset(request.POST or None,  queryset=PRIJobVacancyJobResponsibilitiesInfo.objects.none())

            if form1.is_valid() and form2.is_valid() and form3.is_valid():

                try:
                    instance_form1 = form1.save(commit=False)
                    instance_form1.client_request = requests_post
                    instance_form1.save() 
                    instance_form1.secure_id = encrypt_key(instance_form1.id)
                    instance_form1.save()

                    instance_form2 = form2.save(commit=False)
                    for form in instance_form2:
                        form.job_vacancy_jq = instance_form1
                        form.save()

                    instance_form3 = form3.save(commit=False)

                    for form in instance_form3:
                        form.job_vacancy_jr = instance_form1
                        form.save()

                    data['form_is_valid'] = True
                except IntegrityError:
                    data['form_is_valid'] = False
                    messages.error(request, 'Record Already Exists')

            else:
                data['form_is_valid'] = False
        elif request.method == 'GET':
            form1 = PRIJobVacancyForm(request.GET or None)
            form2 = PRIJobVacancyJobQualificationsFormset(request.GET or None, queryset=PRIJobVacancyJobQualificationsInfo.objects.none())
            form3 = PRIJobVacancyJobResponsibilitiesFormset(request.GET or None, queryset=PRIJobVacancyJobResponsibilitiesInfo.objects.none())
            # form2 = PRIJobVacancyJobQualificationsFormset(request.GET or None, )
            # form3 = PRIJobVacancyJobResponsibilitiesFormset(request.GET or None,)

        context = {
            'record': requests_post,
            'form1': form1,
            'form2': form2,
            'form3': form3,
            'policy': policy,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()

def pri_admin_edit_jobs(request, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_jobs/pri_admin_edit_jobs.html'
    job = get_object_or_404(PRIJobVacancyInfo, id=id)

    user = User.objects.all().filter(username=request.user)     
    if user:   
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_edit_job_vacancy:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))

    PRIJobVacancyJobQualificationsFormset = inlineformset_factory(PRIJobVacancyInfo, PRIJobVacancyJobQualificationsInfo, form=PRIJobVacancyJobQualificationsForm, extra=1, can_delete=True)
    PRIJobVacancyJobResponsibilitiesFormset = inlineformset_factory(PRIJobVacancyInfo, PRIJobVacancyJobResponsibilitiesInfo,  form=PRIJobVacancyJobResponsibilitiesForm, extra=1, can_delete=True)


    if request.is_ajax():
        if request.method == 'POST':
            form1 = PRIJobVacancyForm(request.POST, instance=job)
            form2 = PRIJobVacancyJobQualificationsFormset(request.POST or None, instance=job)
            form3 = PRIJobVacancyJobResponsibilitiesFormset(request.POST or None, instance=job)


            if form1.is_valid() and form2.is_valid() and form3.is_valid():

                form1.save()

                instance_form2 = form2.save(commit=False)
                # for instance in instance_form2:
                #     instance.save()
                form2.save()

                instance_form3 = form3.save(commit=False)
                # for instance in instance_form3:
                #     instance.save()
                form3.save()

                # re query after update
                new_query, count = pri_admin_jobs_filtered_query(request)

                context = {
                    'query': new_query,
                    'count': count,
                    'policy': policy,
                }

                data['table_records_jobs'] = render_to_string('pri_admin/tables/pri_admin_jobs_table.html', context)
                data['pagination_jobs'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

                data['form_is_valid'] = True

            else:
                data['form_is_valid'] = False


        elif request.method == 'GET':
            form1 = PRIJobVacancyForm(request.GET or None, instance=job)
            form2 = PRIJobVacancyJobQualificationsFormset(request.GET or None, instance=job)
            form3 = PRIJobVacancyJobResponsibilitiesFormset(request.GET or None, instance=job)

        context = {
            'job': job,
            'form1': form1,
            'form2': form2,
            'form3': form3,
            'policy': policy,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)

    else:
        raise Http404()

def pri_admin_delete_jobs(request, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_jobs/pri_admin_delete_jobs.html'
    job = get_object_or_404(PRIJobVacancyInfo, id=id)

    user = User.objects.all().filter(username=request.user)     
    if user:   
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_delete_job_vacancy:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))

    if request.is_ajax():
        if request.method == 'POST':
            job.delete()

            new_query, count = pri_admin_jobs_filtered_query(request)

            context = {
                'query': new_query,
                'count': count,
                'policy': policy,
            }

            data['table_records_jobs'] = render_to_string('pri_admin/tables/pri_admin_jobs_table.html', context)
            data['pagination_jobs'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

            data['form_is_valid'] = True

        elif request.method == 'GET':
            context = {
                'job': job,
            }
            data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()

def pri_admin_job_details(request, key):
    template_name = "pri_admin/pri_admin_jobs/pri_admin_job_details.html"

    job_vacancies = get_object_or_404(PRIJobVacancyInfo, id=decrypt_key(key))  

    user = User.objects.all().filter(username=request.user)

    print(job_vacancies)

    print('----------->', request.user)

    try:
        profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))
    except PRIUserInfo.DoesNotExist:
        raise Http404()
    # job hist, edu, skill
    applicant_list = PRIApplicantProfileInfo.objects.all().distinct() 
    
    job_vacancy_matching = [job_vacancies.job_title, job_vacancies.job_category,job_vacancies.job_specialization,job_vacancies.job_minimum_experience,str(job_vacancies.job_salary) ]
    
    applicant_category_matching = []

    for applicant in applicant_list:
        applicant_category = []
        applicant_education= []
        applicant_skills = []
        for empHistory in applicant.employment_applicant_fk.all():
            history_company = empHistory.company
            history_position = empHistory.position
            history_basic_salary_pay = empHistory.basic_salary_pay
            applicant_category.append(history_company)
            applicant_category.append(history_position)
            applicant_category.append(str(history_basic_salary_pay)) 

        for educationAtt in applicant.educational_applicant_fk.all():
            education_college = educationAtt.college
            education_college_course = educationAtt.college_course
            education_vocational = educationAtt.vocational
            education_vocational_course = educationAtt.vocational_course
            education_highschool = educationAtt.highschool
            education_elementary = educationAtt.elementary
            education_special_skills = educationAtt.special_skills
            applicant_education.append(education_college)
            applicant_education.append(education_college_course)
            applicant_education.append(education_vocational)
            applicant_education.append(education_vocational_course)
            applicant_education.append(education_highschool)
            applicant_education.append(education_elementary)
            applicant_skills.append(education_special_skills) 
        
        # For Education
        educ_counter = 0 
        print('-------------->', applicant_education)
        for item in job_vacancy_matching:
            match = any([m for m in applicant_education if item in m])
            if match:
                educ_counter += 1
                print('=>', match)  

        # for employment history
        emp_his_counter = 0
        for item in job_vacancy_matching:
            match = any(m for m in applicant_category if item in m)
            if match:
                emp_his_counter += 1 

        skill_counter = 0
        for item in job_vacancy_matching:
            match = any(m for m in applicant_skills if item in m)
            if match:
                skill_counter += 1 
        # Round 2 decimal places only
        total_ranking = "{}%".format(round((educ_counter * 0.30) + (emp_his_counter * 0.30) + (skill_counter * 0.40),2))
        
 
        default_image = '/media/images/Administrator Male_48px.png'    
        default_file = "No File"

        if applicant.image:
            default_image = applicant.image.url
        
        if applicant.cv:
            default_file = applicant.cv.url

        applicant_category_matching.append([job_vacancies.id, default_image, str(applicant.lname +" "+ applicant.fname +", "+ applicant.mname), str(default_file), applicant.user.email, applicant.mobile_no, applicant.secure_key_id ,total_ranking]) 
  
    
    #test = reduce(operator.or_, (Q(x__icontains=x) for x in applicant_list))
    
    """
    list_of_data = ["asdasdas","qweqwe","zxczxc"]
    another_list = ['a','x','c','e','f']

    # matchers = ['abc','def']
    # matching = [s for s in my_list if any(xs in s for xs in matchers)]
    count=0
    for x in another_list:
        # for s in list_of_data:
        #     if x in s:
        #         print(True)
        #         print(x, '->',s)
        y = any([s for s in list_of_data if x in s])
        if y:
            count+=1
        print(y)

    print(count) 

    """ 
    # for applicant in applicant_list:
    #     test = PRIApplicantProfileInfo.objects.get(id=applicant.id)
    #     print(applicant, ' = ',test, '->' ,applicant.id)

    context = {
        'job_vacancies': job_vacancies,
        'applicants':sort_applicants(applicant_category_matching),
    }

    return render(request, template_name, context)

def sort_applicants(applicant_list):
    length_list = len(applicant_list) - 1
    return(sorted(applicant_list, key=lambda x:x[length_list], reverse=True))

"""
-----------------------------------------------------------------------
**************************GENERAL FUNCTIONS****************************
-----------------------------------------------------------------------
"""

def pri_admin_jobs_database_query(searched_text, sortBy, column):
    if sortBy is None or column is None:
        sortable = '-id'
    else:
        if sortBy.lower() == 'Ascending'.lower():
            sortable = column.lower()
        elif sortBy.lower() == 'Descending'.lower():
            sortable = '-' + column.lower()

    if searched_text:
        query = PRIJobVacancyInfo.objects.filter(
            Q(client_request__requested_client__client_company_name__icontains=searched_text) |
            Q(job_title__icontains=searched_text) |
            Q(job_specialization__icontains=searched_text) |
            Q(job_minimum_experience__icontains=searched_text) |
            Q(job_salary__icontains=searched_text) |
            Q(job_deadline__icontains=searched_text)
        ).distinct().order_by(sortable)
    else:
        query = PRIJobVacancyInfo.objects.all().order_by(sortable)

    return query

def pri_admin_jobs_page_pagination(query, limit, current_page=1):
    if current_page is None:
        current_page = 1

    paginator = Paginator(query, limit)

    try:
        new_query = paginator.page(current_page)
    except PageNotAnInteger:
        print('Error: ', PageNotAnInteger)
        new_query = paginator.page(1)
    except EmptyPage:
        print('Error: ', EmptyPage)
        new_query = paginator.page(paginator.num_pages)

    return new_query

def pri_admin_jobs_filtered_query(request):
    searched_text = request.session.get('search_text') if request.session.get('search_text') is not None else False
    sortBy = request.session.get('sortBy') if request.session.get('sortBy') is not None else 'Descending'
    column = request.session.get('column') if request.session.get('column') is not None else 'id'
    limit = request.session.get('limit') if request.session.get('limit') is not None else 10
    current_page = request.session.get('current_page') if request.session.get('current_page') is not None else 1

    print('SEARCHED: ', searched_text)
    print('SORT BY: ', sortBy)
    print('COLUMN: ', column.lower().replace(" ", "_"))
    print('LIMIT: ', limit)
    print('CURRENT PAGE: ', current_page)

    query = pri_admin_jobs_database_query(searched_text, sortBy, column.lower().replace(" ", "_"))
    count = query.count()

    new_query = pri_admin_jobs_page_pagination(query, limit, current_page)

    return new_query, count

def pri_admin_jobs_search_filter(request):
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['search_text'] = request.GET.get('search_text')

        new_query, count = pri_admin_jobs_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': user_policy(request),
        }

        data['table_records_jobs'] = render_to_string('pri_admin/tables/pri_admin_jobs_table.html', context)
        data['pagination_jobs'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

def pri_admin_jobs_record_limiter(request):
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['limit'] = request.GET.get('limit')

        new_query, count = pri_admin_jobs_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': user_policy(request),
        }

        data['table_records_jobs'] = render_to_string('pri_admin/tables/pri_admin_jobs_table.html', context)
        data['pagination_jobs'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

def pri_admin_jobs_sort_records(request):
    if request.method == 'GET' and request.is_ajax():
        data = dict()

        request.session['sortBy'] = request.GET.get('sortBy')
        request.session['column'] = request.GET.get('column')

        new_query, count = pri_admin_jobs_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': user_policy(request),
        }

        data['table_records_jobs'] = render_to_string('pri_admin/tables/pri_admin_jobs_table.html', context)
        data['pagination_jobs'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

def pri_admin_jobs_paging(request):
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['current_page'] = request.GET.get('page')

        new_query, count = pri_admin_jobs_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': user_policy(request),
        }

        data['table_records_jobs'] = render_to_string('pri_admin/tables/pri_admin_jobs_table.html', context)
        data['pagination_jobs'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

"""
=======================================================================
***************************END PRI Jobs******************************** 
=======================================================================
"""


"""
For PRI Application Form
"""


class PriApplicationForm(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_application_form_page.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriApplicationForm, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)

        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))
        except PRIUserInfo.DoesNotExist:
            raise Http404()

        context.update({
            'user': user,
            'profile': profile
        })

        return context


"""
For PRI Applicants
"""


class PriApplicants(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_applicants_page.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriApplicants, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)
        applicants = PRIApplicantProfileInfo.objects.all().order_by('-id').distinct()
      
        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))
              # Policy
            pri_user = PRIUserInfo.objects.all().get(user__in=user)
            policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
            if not policy.pri_can_view_applicants_page:
                raise Http404()         

        except PRIUserInfo.DoesNotExist:
            raise Http404()

        context.update({
            'user': user,
            'profile': profile,
            'applicants': applicants,
            'policy': policy,
        })

        return context

def pri_admin_delete_applicants(request, key):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicants/pri_admin_delete_applicants.html'
    id = decrypt_key(key)
    applicant_profile = get_object_or_404(PRIApplicantProfileInfo, id=id)
    user = User.objects.all().get(applicant_user_fk=applicant_profile)
    applicants = PRIApplicantProfileInfo.objects.all().order_by('-id').distinct()

    
    myuser = User.objects.all().filter(username=request.user)     
    if myuser:   
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=myuser)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_delete_applicants:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))

    if request.is_ajax():
        if request.method == 'POST':
            user.delete()

            context = {
                'applicants': applicants,
                'policy': policy,
            }
            data['table_records_applicants'] = render_to_string('pri_admin/tables/pri_admin_applicants_table.html', context)

            data['form_is_valid'] = True
            
        elif request.method == 'GET':
            context = {
                'user': user,
                'applicant_profile': applicant_profile,
            }
            data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


"""
For PRI Schedules
"""

class PriSchedules(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_schedules.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriSchedules, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)
        job_requests = PRIApplicantJobRequestInfo.objects.all().order_by('-id').distinct()

        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))
        except PRIUserInfo.DoesNotExist:
            raise Http404()
 

        context.update({
            'user': user,
            'profile': profile,
            'job_requests': job_requests,
        })

        return context


def pri_schedules_toggle_exam(request, id):
    data = dict()
    job_applicant_request = get_object_or_404(PRIApplicantJobRequestInfo, id=id)
    if request.is_ajax():
        if request.method == 'GET':
            if job_applicant_request.take_exam: 
                job_applicant_request.take_exam = False
                job_applicant_request.save()
            else:
                job_applicant_request.take_exam = True
                job_applicant_request.save()
            job_requests = PRIApplicantJobRequestInfo.objects.all().order_by('-id').distinct()
            context = {
                'job_requests': job_requests,
            }

            data['table_records_schedule'] = render_to_string('pri_admin/tables/pri_admin_schedules_tables.html', context)

            return JsonResponse(data)
    else:
        raise Http404()
"""
=======================================================
******************PRI REQUEST PAGE*********************
=======================================================
"""


"""
For PRI Request
"""

class PriRequest(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_requests.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriRequest, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)  

        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))
            # Policy
            pri_user = PRIUserInfo.objects.all().get(user__in=user)
            policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
            if not policy.pri_can_view_request_page:
                raise Http404()         
        except PRIUserInfo.DoesNotExist:
            raise Http404()

        for key, value in self.request.session.items():
            print('{} -> {}'.format(key, value))

        if 'current_page' in self.request.session:
            del self.request.session['current_page']
            print('current_page => DELETED!')
        else:
            print('current_page is not existing yet!')

        if 'search_text' in self.request.session:
            del self.request.session['search_text']
            print('searched_text => DELETED!')
        else:
            print('searched_text is not existing yet')

        if 'sortBy' in self.request.session:
            del self.request.session['sortBy']
            print('sortBy => DELETED!')
        else:
            print('sortBy is not existing yet')

        if 'column' in self.request.session:
            del self.request.session['column']
            print('column => DELETED')
        else:
            print('column is not existing yet')

        if 'limit' in self.request.session:
            del self.request.session['limit']
            print('limit => DELETED')
        else:
            print('limit is not existing yet')

        query = PRIRequestInfo.objects.all().order_by('-id')
        count = query.count()
        query = pri_admin_requests_page_pagination(query, 10, 1)

        columns = []

        for field in PRIRequestInfo._meta.get_fields()[1:]:
            col = field.get_attname_column()[1]
            columns.append(col.replace("_", " "))

        columns.remove("content")
        columns.remove("seen by admin")

        context.update({
            'user': user,
            'profile': profile,
            'query': query,
            'count': count,
            'columns': columns,
            'policy': policy,
        })

        return context

"""
-------------------------------------------------------
****************PRI REQUEST FUNCTIONS******************
-------------------------------------------------------
"""
def pri_admin_create_requests(request):
    data = dict()
    template_name = 'pri_admin/pri_admin_requests/pri_admin_create_requests.html'

    user = User.objects.all().filter(username=request.user)        
    if user:
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_add_request:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))


    if request.is_ajax():
        if request.method == 'POST':
            form = PRIRequestForm(request.POST)
            if form.is_valid():

                form.save()

                new_query, count = pri_admin_requests_filtered_query(request)

                context = {
                    'query': new_query,
                    'count': count,
                    'policy': policy,
                }

                data['table_records_requests'] = render_to_string('pri_admin/tables/pri_admin_requests_table.html', context)
                data['pagination_requests'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        elif request.method == 'GET':
            form = PRIRequestForm()

        context = {
            'form': form,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()

def pri_admin_edit_requests(request, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_requests/pri_admin_edit_requests.html'
    record = get_object_or_404(PRIRequestInfo, id=id)

    user = User.objects.all().filter(username=request.user)        
    if user:
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_edit_request:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))


    if request.is_ajax():
        if request.method == 'POST':
            form = PRIRequestForm(request.POST, instance=record)
            if form.is_valid():

                instance = form.save(commit=False)
                instance.seen_by_admin = False
                instance.save()

                new_query, count = pri_admin_requests_filtered_query(request)

                context = {
                    'query': new_query,
                    'count': count,
                    'policy': policy,
                }

                data['table_records_requests'] = render_to_string('pri_admin/tables/pri_admin_requests_table.html', context)
                data['pagination_requests'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        elif request.method == 'GET':
            form = PRIRequestForm(instance=record)

        context = {
            'form': form,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()

def pri_admin_delete_requests(request, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_requests/pri_admin_delete_requests.html'
    record = get_object_or_404(PRIRequestInfo, id=id)

    user = User.objects.all().filter(username=request.user)    
    if user:    
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_delete_request:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))


    if request.is_ajax():
        if request.method == 'POST':

            record.delete()

            new_query, count = pri_admin_requests_filtered_query(request)

            context = {
                'query': new_query,
                'count': count,
                'policy': policy,
            }

            data['table_records_requests'] = render_to_string('pri_admin/tables/pri_admin_requests_table.html', context)
            data['pagination_requests'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

            data['form_is_valid'] = True

        elif request.method == 'GET':

            context = {
                'record': record,
            }

            data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)

    else:
        raise Http404()

def pri_admin_view_requests(request, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_requests/pri_admin_view_requests.html'
    record = get_object_or_404(PRIRequestInfo, id=id)

    user = User.objects.all().filter(username=request.user)        
    if user:
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        if not policy.pri_can_view_request:
            raise Http404()         
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard"))


    if request.is_ajax():
        if request.method == 'GET':

            record.seen_by_admin = True
            record.save()

            new_query, count = pri_admin_requests_filtered_query(request)

            context = {
                'query': new_query,
                'count': count,
                'record': record,
                'policy': policy,
            }

            data['table_records_requests'] = render_to_string('pri_admin/tables/pri_admin_requests_table.html', context)
            data['pagination_requests'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

            data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()

"""
------------------------------------------------------- 
"""

def user_policy(request): 
    user = User.objects.all().filter(username=request.user)     
    if user:   
        # Policy
        pri_user = PRIUserInfo.objects.all().get(user__in=user)
        policy = PRIUserPermission.objects.all().get(pri_user=pri_user) 
        return policy
    else:
        return HttpResponseRedirect(reverse_lazy("pri:dashboard")) 


def pri_admin_requests_database_query(searched_text, sortBy, column):
    if sortBy is None or column is None:
        sortable = '-id'
    else:
        if sortBy.lower() == 'Ascending'.lower():
            sortable = column.lower()
        elif sortBy.lower() == 'Descending'.lower():
            sortable = '-' + column.lower()

    if searched_text:
        query = PRIRequestInfo.objects.filter(
            Q(requested_client__client_company_name__icontains=searched_text) |
            Q(status__icontains=searched_text) |
            Q(content__icontains=searched_text) |
            Q(data_requested__icontains=searched_text)
        ).distinct().order_by(sortable)
    else:
        query = PRIRequestInfo.objects.all().order_by(sortable)

    return query

def pri_admin_requests_page_pagination(query, limit, current_page=1):
    if current_page is None:
        current_page = 1

    paginator = Paginator(query, limit)

    try:
        new_query = paginator.page(current_page)
    except PageNotAnInteger:
        print('Error: ', PageNotAnInteger)
        new_query = paginator.page(1)
    except EmptyPage:
        print('Error: ', EmptyPage)
        new_query = paginator.page(paginator.num_pages)

    return new_query

def pri_admin_requests_filtered_query(request):
    searched_text = request.session.get('search_text') if request.session.get('search_text') is not None else False
    sortBy = request.session.get('sortBy') if request.session.get('sortBy') is not None else 'Descending'
    column = request.session.get('column') if request.session.get('column') is not None else 'id'
    limit = request.session.get('limit') if request.session.get('limit') is not None else 10
    current_page = request.session.get('current_page') if request.session.get('current_page') is not None else 1

    print('SEARCHED: ', searched_text)
    print('SORT BY: ', sortBy)
    print('COLUMN: ', column.lower().replace(" ", "_"))
    print('LIMIT: ', limit)
    print('CURRENT PAGE: ', current_page)

    query = pri_admin_requests_database_query(searched_text, sortBy, column.lower().replace(" ", "_"))
    count = query.count()

    new_query = pri_admin_requests_page_pagination(query, limit, current_page)

    return new_query, count

def pri_admin_requests_search_filter(request):
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['search_text'] = request.GET.get('search_text')

        new_query, count = pri_admin_requests_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': user_policy(request),
        }

        data['table_records_requests'] = render_to_string('pri_admin/tables/pri_admin_requests_table.html', context)
        data['pagination_requests'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

def pri_admin_requests_record_limiter(request):
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['limit'] = request.GET.get('limit')

        new_query, count = pri_admin_requests_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': user_policy(request),
        }

        data['table_records_requests'] = render_to_string('pri_admin/tables/pri_admin_requests_table.html', context)
        data['pagination_requests'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

def pri_admin_requests_sort_records(request):
    if request.method == 'GET' and request.is_ajax():
        data = dict()

        request.session['sortBy'] = request.GET.get('sortBy')
        request.session['column'] = request.GET.get('column')

        new_query, count = pri_admin_requests_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': user_policy(request),
        }

        data['table_records_requests'] = render_to_string('pri_admin/tables/pri_admin_requests_table.html', context)
        data['pagination_requests'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

def pri_admin_requests_paging(request):
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['current_page'] = request.GET.get('page')

        new_query, count = pri_admin_requests_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
            'policy': user_policy(request),
        }

        data['table_records_requests'] = render_to_string('pri_admin/tables/pri_admin_requests_table.html', context)
        data['pagination_requests'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()



"""
=======================================================
**************END PRI REQUEST FUNCTIONS****************
=======================================================
"""

"""
For PRI Examinations
"""

class PriExamninations(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_examinations.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriExamninations, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)

        exam = ExaminationInfo.objects.all().filter(exam_type='Quiz')

        question = QuestionInfo.objects.all().filter(examination__in=exam)
        choices = ChoicesInfo.objects.all().filter(question__in=question)


        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(Q(level='Administrator') | Q(level='Sub Admin') | Q(level='Employee')))
        except PRIUserInfo.DoesNotExist:
            raise Http404()

        
        queryset = serializers.serialize('json', question)

        print(queryset)



        context.update({
            'user': user,
            'profile': profile,
            'exam': exam,
            'question': question,
            'choices': choices,
        })

        return context


#test report
#https://openpyxl.readthedocs.io/en/stable/

def export_users_csv(request):
    response = HttpResponse(content_type="text/csv")
    response['Content-Disposition'] = 'attachment; filename="users.csv"'

    writer = csv.writer(response)
    writer.writerow(['Username', 'First name', 'Last name', 'Email address'])

    users = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')
    for user in users:
        writer.writerow(user)


    return response


def export_users_xls(request):
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename={date}-users.xls'.format(date=datetime.now().strftime('%Y-%m-%d'),)

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Users')

    # Sheet header, first row

    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Username', 'First name', 'Last name', 'Email address', ]

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows

    font_style = xlwt.XFStyle()

    rows = User.objects.all().values_list('username', 'first_name', 'last_name', 'email')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)

    wb.save(response)

    return response

def export_users_xlsx(request):
    #https://openpyxl.readthedocs.io/en/stable/tutorial.html
    user = User.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename={date}-users.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

    workbook = Workbook()

    ws = workbook.active

    ws = workbook.create_sheet("MySheet")

    ws.title = "Users"

    ws.sheet_properties.tabColor = "1072BA"

    ws.cell(row=4, column=2, value=10)

    for x in range(1, 101):
        for y in range(1, 101):
            ws.cell(row=x, column=y, value=y)

    a1 = ws['A1']
    ft = Font(color=colors.RED, italic=True, size=15)



    # If you want to change the color of a Font, you need to reassign it::

    thin = Side(border_style="thin", color="007BFF")
    thick = Side(border_style="thick", color="2C6D04")
    double = Side(border_style="double", color="ff0000")

    a1.font = ft  # the change only affects A1
    a1.border = Border(top=double, left=thin, right=thin, bottom=thick)
    # col = ws.column_dimensions['A']
    # col.font = Font(bold=True)
    # row = ws.row_dimensions[1]
    # row.font = Font(underline="single")

    # Merge cells

    ws.merge_cells('B2:F4')

    top_left_cell = ws['B2']
    top_left_cell.value = "My Cell"



    top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
    top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
    top_left_cell.fill = GradientFill(stop=("000000", "FFFFFF"))
    top_left_cell.font = Font(b=True, color="FF0000")
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")


    print(workbook.sheetnames)

    # wb.save('balances.xlsx')
    workbook.save(response)

    return response





"""
reverse() function: It’s similar to the url template tag which use to convert namespaced url to real url pattern.
reverse_lazy() function: This is a reverse() function’s lazy version. It’s prevent to occur error when URLConf is not loaded. Generally we use this function in case below:
providing a reversed URL as the url attribute of a generic class-based view.
providing a reversed URL to a decorator (such as the login_url argument for the django.contrib.auth.decorators.permission_required() decorator)
providing a reversed URL as a default value for a parameter in a function’s signature.
"""
def pri_user_login(request):

    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)
        current_user = User.objects.all().filter(username=user) 
        if current_user:
            current_user = User.objects.all().get(username=user) 
        if user:
            if user.is_active and user.is_staff and user.is_superuser:
                request.session.set_expiry(300)
                login(request, user)
                return HttpResponseRedirect(reverse_lazy("pri:dashboard")) 
            elif user.is_active and user.is_staff and not user.is_superuser: 
                try:
                    pri_user = current_user.user_fk.level
                    if pri_user.casefold() == "employee" or pri_user.casefold() == "sub admin":
                        request.session.set_expiry(300)
                        login(request, user)
                        return HttpResponseRedirect(reverse_lazy("pri:dashboard")) 
                    elif pri_user.casefold() == "client":
                        request.session.set_expiry(300)
                        login(request, user)           
                        return HttpResponseRedirect(reverse_lazy("client_side_home_page"))
                except User.user_fk.RelatedObjectDoesNotExist: 
                    return HttpResponseRedirect(reverse_lazy("user_login"))
            elif user.is_active and not user.is_staff and not user.is_superuser: 
                request.session.set_expiry(300)
                login(request, user)
                return HttpResponseRedirect(reverse_lazy("applicant_side_job_posts_all"))
            else:
                messages.warning(request, "Account Not Active")
        else:
            messages.error(request, "Invalid Account")

    return render(request, "registration/login.html", {})

"""
==========================================================
***********************PRI ADMIN**************************
==========================================================
"""
"""
For PRI Admin
"""


class PriAdmin(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_admin_page.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PriAdmin, self).get_context_data(*args, **kwargs)
        user = User.objects.all().filter(username=self.request.user)     

        try:
            profile = PRIUserInfo.objects.all().get(Q(user__in=user) & Q(level='Administrator'))
            # Policy
            pri_user = PRIUserInfo.objects.all().get(user__in=user)
            policy = PRIUserPermission.objects.all().get(pri_user=pri_user)
        except PRIUserInfo.DoesNotExist:
            raise Http404()

        for key, value in self.request.session.items():
            print('{} -> {}'.format(key, value))

        if 'current_page' in self.request.session:
            del self.request.session['current_page']
            print('current_page => DELETED!')
        else:
            print('current_page is not existing yet!')

        if 'search_text' in self.request.session:
            del self.request.session['search_text']
            print('searched_text => DELETED!')
        else:
            print('searched_text is not existing yet')

        if 'sortBy' in self.request.session:
            del self.request.session['sortBy']
            print('sortBy => DELETED!')
        else:
            print('sortBy is not existing yet')

        if 'column' in self.request.session:
            del self.request.session['column']
            print('column => DELETED')
        else:
            print('column is not existing yet')

        if 'limit' in self.request.session:
            del self.request.session['limit']
            print('limit => DELETED')
        else:
            print('limit is not existing yet')

        query = PRIUserInfo.objects.all().order_by('-id')
        count = query.count()
        query = pri_admin_users_page_pagination(query, 10, 1)

        columns = []

        for field in PRIUserInfo._meta.get_fields()[2:]:
            col = field.get_attname_column()[1]
            columns.append(col.replace("_", " "))

        columns.remove("user id")
        columns.remove("age")
        columns.remove("gender")
        columns.remove("address")
        columns.remove("image")
        columns.remove("date added")

        context.update({
            'user': user,
            'profile': profile,
            'query': query,
            'count': count,
            'columns': columns,
            'policy': policy,
        })

        return context

def pri_admin_apply_policy(request, id):
    data = dict()
    template_name = "pri_admin/pri_admin_users/pri_admin_apply_policy.html"
    pri_user = get_object_or_404(PRIUserInfo, id=id) 
    
    if request.is_ajax():
        if request.method == "POST":
            form = PRIUserPermissionForm(request.POST or None) 
            if form.is_valid(): 
                policy, created = PRIUserPermission.objects.update_or_create(pri_user=pri_user, defaults={**form.cleaned_data})
                if created:
                    print("Already Exists!")
                else:
                    print("New Record!")
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        elif request.method == "GET":
            try:
                pri_policy = PRIUserPermission.objects.all().get(pri_user=pri_user)  
                form = PRIUserPermissionForm(request.GET or None, instance=pri_policy) 
            except PRIUserPermission.DoesNotExist:
                form = PRIUserPermissionForm(request.GET or None) 
        context = {
            'form': form,
            'record': pri_user,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)

    else:
        raise Http404()


"""
PRI Admin Users Functions
"""


def pri_admin_users_database_query(searched_text, sortBy, column):

    if sortBy is None or column is None:
        sortable = '-id'
    else:
        if sortBy.lower() == 'Ascending'.lower():
            sortable = column.lower()
        elif sortBy.lower() == 'Descending'.lower():
            sortable = '-' + column.lower()

    if searched_text:
        query = PRIUserInfo.objects.filter(
            Q(first_name__icontains=searched_text) |
            Q(middle_name__icontains=searched_text) |
            Q(last_name__icontains=searched_text) |
            Q(contact__icontains=searched_text) |
            Q(address__icontains=searched_text) |
            Q(position__icontains=searched_text) |
            Q(department__icontains=searched_text) |
            Q(level__icontains=searched_text) |
            Q(user__email__icontains=searched_text)
        ).distinct().order_by(sortable)
    else:
        query = PRIUserInfo.objects.all().order_by(sortable)

    return query


def pri_admin_users_page_pagination(query, limit, current_page=1):
    if current_page is None:
        current_page = 1

    paginator = Paginator(query, limit)

    try:
        new_query = paginator.page(current_page)
    except PageNotAnInteger:
        print('Error: ', PageNotAnInteger)
        new_query = paginator.page(1)
    except EmptyPage:
        print('Error: ', EmptyPage)
        new_query = paginator.page(paginator.num_pages)

    return new_query


def pri_admin_users_filtered_query(request):
    searched_text = request.session.get('search_text') if request.session.get('search_text') is not None else False
    sortBy = request.session.get('sortBy') if request.session.get('sortBy') is not None else 'Descending'
    column = request.session.get('column') if request.session.get('column') is not None else 'id'
    limit = request.session.get('limit') if request.session.get('limit') is not None else 10
    current_page = request.session.get('current_page') if request.session.get('current_page') is not None else 1

    print('SEARCHED: ', searched_text)
    print('SORT BY: ', sortBy)
    print('COLUMN: ', column.lower().replace(" ", "_"))
    print('LIMIT: ', limit)
    print('CURRENT PAGE: ', current_page)

    query = pri_admin_users_database_query(searched_text, sortBy, column.lower().replace(" ", "_"))
    count = query.count()

    new_query = pri_admin_users_page_pagination(query, limit, current_page)

    return new_query, count


def pri_admin_users_search_filter(request):
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['search_text'] = request.GET.get('search_text')

        new_query, count = pri_admin_users_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
        }

        data['table_records_admin'] = render_to_string('pri_admin/tables/pri_admin_users_table.html', context)
        data['pagination_admin'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()


def pri_admin_users_record_limiter(request):
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['limit'] = request.GET.get('limit')

        new_query, count = pri_admin_users_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
        }

        data['table_records_admin'] = render_to_string('pri_admin/tables/pri_admin_users_table.html', context)
        data['pagination_admin'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()


def pri_admin_users_sort_records(request):
    if request.method == 'GET' and request.is_ajax():
        data = dict()

        request.session['sortBy'] = request.GET.get('sortBy')
        request.session['column'] = request.GET.get('column')

        new_query, count = pri_admin_users_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
        }

        data['table_records_admin'] = render_to_string('pri_admin/tables/pri_admin_users_table.html', context)
        data['pagination_admin'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)


        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()


def pri_admin_users_paging(request):
    if request.method == 'GET' and request.is_ajax():

        data = dict()

        request.session['current_page'] = request.GET.get('page')

        new_query, count = pri_admin_users_filtered_query(request)

        context = {
            'query': new_query,
            'count': count,
        }

        data['table_records_admin'] = render_to_string('pri_admin/tables/pri_admin_users_table.html', context)
        data['pagination_admin'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

        return JsonResponse(data)
    else:
        # value error when directly to url not ajax request
        raise Http404()

"""
PRI Admin users 
"""

def pri_admin_users_create_account_form(request):
    data = dict()
    template_name = 'pri_admin/pri_admin_users/pri_admin_users_create_account.html'

    if request.is_ajax():
        if request.method == 'POST':
            profile_form = PRIUserProfileRegistrationForm(request.POST, request.FILES or None)
            account_form = PRIUserAccountRegistrationForm(request.POST)

            if profile_form.is_valid() and account_form.is_valid():

                acc_frm = account_form.save(commit=False)
                acc_frm.is_active = False
                acc_frm.is_staff = False
                acc_frm.is_superuser = False
                acc_frm.save()

                prf_frm = profile_form.save(commit=False)
                prf_frm.user = acc_frm

                if 'image' in request.FILES:
                    prf_frm.image = request.FILES['image']

                prf_frm.save()

                new_query, count = pri_admin_users_filtered_query(request)

                context = {
                    'query': new_query,
                    'count': count,
                }
                # create default permissiob
                policy, created = PRIUserPermission.objects.update_or_create(pri_user=prf_frm)
                if created:
                    print("Already Exists!")
                else:
                    print("New Record!")

                data['table_records_admin'] = render_to_string('pri_admin/tables/pri_admin_users_table.html', context)
                data['pagination_admin'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

                data['form_is_valid'] = True

            else:
                data['form_is_valid'] = False
                messages.error(request, 'Form Error!')

        elif request.method == 'GET':

            profile_form = PRIUserProfileRegistrationForm()
            account_form = PRIUserAccountRegistrationForm()

        context = {
            'form': profile_form,
            'account_form': account_form,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)

    else:
        raise Http404()



"""
==========================================================
PRI Admin Update Profile
==========================================================
"""

def pri_admin_update_profile(request, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_users/pri_admin_update_profile.html'
    profile = get_object_or_404(PRIUserInfo, id=id)
    user = User.objects.all().get(user_fk=profile)

    if request.is_ajax():
        if request.method == 'POST':

            form = PRIUserProfileRegistrationForm(request.POST, request.FILES or None, instance=profile)

            if form.is_valid():
                instance = form.save(commit=False)

                if 'image' in request.FILES:
                    instance.image = request.FILES['image']
                else:
                    print("NO IMAGE DETECTED!")

                instance.save()

                new_query, count = pri_admin_users_filtered_query(request)

                context = {
                    'query': new_query,
                    'count': count,
                }

                data['table_records_admin'] = render_to_string('pri_admin/tables/pri_admin_users_table.html', context)
                data['pagination_admin'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

                data['form_is_valid'] = True

            else:
                data['form_is_valid'] = False

        elif request.method == 'GET':
            form = PRIUserProfileRegistrationForm(request.FILES or None, instance=profile)

        context = {
            'profile': profile,
            'form': form,
            'user': user,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)

    else:
        raise Http404()


def pri_admin_users_delete_form(request, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_users/pri_admin_users_delete_account.html'
    profile = get_object_or_404(PRIUserInfo, id=id)
    user = get_object_or_404(User, user_fk=profile)

    if request.is_ajax():
        if request.method == "POST":

            user.delete()

            new_query, count = pri_admin_users_filtered_query(request)

            context = {
                'query': new_query,
                'count': count,
            }

            data['table_records_admin'] = render_to_string('pri_admin/tables/pri_admin_users_table.html', context)
            data['pagination_admin'] = render_to_string('pri_admin/paginators/pri_admin_users_pagination.html', context)

            data['form_is_valid'] = True

        elif request.method == "GET":
            context = {
                'profile': profile,
                'user': user,
            }

            data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)

    else:
        raise Http404()

"""
==============================================================================================
****************************APPLICANT SIDE VIEWS / FUNCTIONS**********************************
==============================================================================================
"""

"""
==============================================
********APPLICANT REGISTRATiON PAGE*********
==============================================
"""

def pri_applicant_account_registration_page(request):
    template_name = 'registration/applicant_registration.html'

    if request.method == 'POST':
        form = PRIUserAccountRegistrationForm(data=request.POST)    
        if form.is_valid():
            user_form = form.save(commit=False)
            user_form.is_active = True 
            user_form.save()

            return HttpResponseRedirect(reverse('user_login'))
        else:
            messages.error(request, 'Form Error!')

        
    elif request.method == 'GET':
        form = PRIUserAccountRegistrationForm(request.GET or None)

    context = {
        'form': form,
    }

    return render(request, template_name, context)

def pri_applicant_registration_page(request, slug):
    #username = User.objects.all().get(username=request.user.username)
    try:
        #user = get_object_or_404(User, username=username)
        user = User.objects.all().get(username=slug) 
        # to be checked
        # if not str(request.user.username) == user.username:
        #     #raise Http404()
        #     return HttpResponseRedirect(reverse_lazy('user_login'))
    except User.DoesNotExist: 
        return HttpResponseRedirect(reverse_lazy('user_login'))
        #raise Http404()

    template_name = 'pri_applicant_side/application_registration.html'

    formApplicantSiblingsFormset = inlineformset_factory(PRIApplicantProfileInfo, PRIApplicantSibilingsInfo, form=PRIApplicantSibilingsForm, extra=1, can_delete=False)
    formApplicantEmploymentHistoryFormset = inlineformset_factory(PRIApplicantProfileInfo, PRIApplicantEmploymentHistoryInfo, form=PRIApplicantEmploymentHistoryForm, extra=1, can_delete=False)
    formApplicantTrainingsFormset = inlineformset_factory(PRIApplicantProfileInfo, PRIApplicantTrainingsInfo, form=PRIApplicantTrainingsForm, extra=1, can_delete=False)
    formApplicantCharacterReferenceFormset = inlineformset_factory(PRIApplicantProfileInfo, PRIApplicantCharacterReferencesInfo, form=PRIApplicantCharacterReferencesForm, extra=1, can_delete=False)
    
    if request.method == 'GET':
        formApplicantProfile=PRIApplicantProfileForm(request.GET or None, request.FILES or None)
        formApplicantSiblings=formApplicantSiblingsFormset(request.GET or None)
        formApplicantEmploymentHistory=formApplicantEmploymentHistoryFormset(request.GET or None)
        formApplicantEducationalAttainment=PRIApplicantEducationalAttainmentForm(request.GET or None)
        formApplicantTrainings=formApplicantTrainingsFormset(request.GET or None)
        formApplicantCharacterReference=formApplicantCharacterReferenceFormset(request.GET or None) 
    elif request.method == 'POST':
        formApplicantProfile=PRIApplicantProfileForm(request.POST or None, request.FILES or None)
        formApplicantSiblings=formApplicantSiblingsFormset(request.POST or None)
        formApplicantEmploymentHistory=formApplicantEmploymentHistoryFormset(request.POST or None)
        formApplicantEducationalAttainment=PRIApplicantEducationalAttainmentForm(request.POST)
        formApplicantTrainings=formApplicantTrainingsFormset(request.POST or None)
        formApplicantCharacterReference=formApplicantCharacterReferenceFormset(request.POST or None) 

        if formApplicantProfile.is_valid() and formApplicantSiblings.is_valid() and formApplicantEmploymentHistory.is_valid() and formApplicantEducationalAttainment.is_valid() and formApplicantTrainings.is_valid() and formApplicantCharacterReference.is_valid():
            # Applicant Profile
            frmAppPro = formApplicantProfile.save(commit=False)
            frmAppPro.user = user
            frmAppPro.save()
            #enecrypt pk and save to different field
            frmAppPro.secure_key_id = encrypt_key(frmAppPro.id)
            frmAppPro.save()
            # Applicant Siblings
            frmAppSib = formApplicantSiblings.save(commit=False)
            for form in frmAppSib:
                form.applicant_siblings = frmAppPro
                form.save()
            # Applicant Employment History
            frmAppEmpHis = formApplicantEmploymentHistory.save(commit=False)
            for form in frmAppEmpHis:
                form.applicant_employment_history = frmAppPro
                form.save()
            # Applicant Educational Attainment
            frmAppEduAtt = formApplicantEducationalAttainment.save(commit=False)
            frmAppEduAtt.applicant_educational_attainment = frmAppPro
            frmAppEduAtt.save()

            # Applicant Trainings
            frmAppTra = formApplicantTrainings.save(commit=False)
            for form in frmAppTra:
                form.applicant_trainings = frmAppPro
                form.save()
            
            # Applicant Character Reference
            frmChaRef = formApplicantCharacterReference.save(commit=False)
            for form in frmChaRef:
                form.applicant_character_references = frmAppPro
                form.save()

            return HttpResponseRedirect(reverse_lazy("applicant_side_job_posts_all"))
        else:
            messages.error(request, "Form Error!")
 
    context = {
        'formApplicantProfile': formApplicantProfile,
        'formApplicantSiblings': formApplicantSiblings,
        'formApplicantEmploymentHistory': formApplicantEmploymentHistory,
        'formApplicantEducationalAttainment': formApplicantEducationalAttainment,
        'formApplicantTrainings': formApplicantTrainings,
        'formApplicantCharacterReference': formApplicantCharacterReference,
    }

    return render(request, template_name, context)


def pri_applicant_edit_profile_page(request, slug):

    template_name = 'pri_applicant_side/applicant_edit_profile.html'

    try:
        #user = get_object_or_404(User, username=username)
        user = User.objects.all().get(username=slug) 
        profile = get_object_or_404(PRIApplicantProfileInfo, user=user)
        education = get_object_or_404(PRIApplicantEducationalAttainmentInfo, applicant_educational_attainment=profile)
        # to be checked
        # if not str(request.user.username) == user.username:
        #     return HttpResponseRedirect(reverse_lazy('user_login'))
        #     #raise Http404()
    except User.DoesNotExist: 
        return HttpResponseRedirect(reverse_lazy('user_login'))
        #raise Http404()


    formApplicantSiblingsFormset = inlineformset_factory(PRIApplicantProfileInfo, PRIApplicantSibilingsInfo, form=PRIApplicantSibilingsForm, extra=1, can_delete=True)
    formApplicantEmploymentHistoryFormset = inlineformset_factory(PRIApplicantProfileInfo, PRIApplicantEmploymentHistoryInfo, form=PRIApplicantEmploymentHistoryForm, extra=1, can_delete=True)
    formApplicantTrainingsFormset = inlineformset_factory(PRIApplicantProfileInfo, PRIApplicantTrainingsInfo, form=PRIApplicantTrainingsForm, extra=1, can_delete=True)
    formApplicantCharacterReferenceFormset = inlineformset_factory(PRIApplicantProfileInfo, PRIApplicantCharacterReferencesInfo, form=PRIApplicantCharacterReferencesForm, extra=1, can_delete=True)

    if request.method == 'GET': 
        formApplicantProfile=PRIApplicantProfileForm(request.GET or None, request.FILES or None, instance=profile)
        formApplicantSiblings=formApplicantSiblingsFormset(request.GET or None, instance=profile)
        formApplicantEmploymentHistory=formApplicantEmploymentHistoryFormset(request.GET or None, instance=profile)
        formApplicantEducationalAttainment=PRIApplicantEducationalAttainmentForm(request.GET or None, instance=education)
        formApplicantTrainings=formApplicantTrainingsFormset(request.GET or None, instance=profile)
        formApplicantCharacterReference=formApplicantCharacterReferenceFormset(request.GET or None, instance=profile) 

    elif request.method == 'POST':
        formApplicantProfile=PRIApplicantProfileForm(request.POST or None, request.FILES or None, instance=profile)
        formApplicantSiblings=formApplicantSiblingsFormset(request.POST or None, instance=profile)
        formApplicantEmploymentHistory=formApplicantEmploymentHistoryFormset(request.POST or None, instance=profile)
        formApplicantEducationalAttainment=PRIApplicantEducationalAttainmentForm(request.POST or None, instance=education)
        formApplicantTrainings=formApplicantTrainingsFormset(request.POST or None, instance=profile)
        formApplicantCharacterReference=formApplicantCharacterReferenceFormset(request.POST or None, instance=profile) 

        if formApplicantProfile.is_valid() and formApplicantSiblings.is_valid() and formApplicantEmploymentHistory.is_valid() and formApplicantEducationalAttainment.is_valid() and formApplicantTrainings.is_valid() and formApplicantCharacterReference.is_valid(): 
        
            # Applicant Profile
            frmAppPro = formApplicantProfile.save(commit=False) 
            frmAppPro.save()
            # Applicant Siblings

            #https://docs.djangoproject.com/es/2.1/topics/forms/#looping-over-hidden-and-visible-fields
            """
            If you’re manually laying out a form in a template, as opposed to relying on Django’s default form layout, 
            you might want to treat <input type="hidden"> fields differently from non-hidden fields. 
            For example, because hidden fields don’t display anything, putting error messages «next to» the field 
            could cause confusion for your users – so errors for those fields should be handled differently.

            Django provides two methods on a form that allow you to loop over the hidden and visible fields 
            independently: hidden_fields() and visible_fields(). Here’s a modification of an earlier example that uses 
            these two methods:

            {# Include the hidden fields #}
            {% if form in formset.forms %}
                {% for hidden in form.hidden_fields %}
                    {{ hidden }}    
                {% endfor %}
            {% endif %}            
            {# Include the visible fields #}
            {% for field in form.visible_fields %}
                <div class="fieldWrapper">
                    {{ field.errors }}
                    {{ field.label_tag }} {{ field }}
                </div>
            {% endfor %}

            This example does not handle any errors in the hidden fields. Usually, an error in a hidden field is a 
            sign of form tampering, since normal form interaction won’t alter them. However, you could easily insert 
            some error displays for those form errors, as well.

            """
            # Hidden id causes error id is required
            for form in formApplicantSiblings.forms: 
                for hidden in form.hidden_fields():                    
                    print(hidden)

            frmAppSib = formApplicantSiblings.save(commit=False) 
            formApplicantSiblings.save() 

            # Applicant Employment History
            frmAppEmpHis = formApplicantEmploymentHistory.save(commit=False)
            formApplicantEmploymentHistory.save()
            # Applicant Educational Attainment
            formApplicantEducationalAttainment.save()
            # Applicant Trainings
            frmAppTra = formApplicantTrainings.save(commit=False)
            formApplicantTrainings.save()
            # Applicant Character Reference
            frmChaRef = formApplicantCharacterReference.save(commit=False)
            formApplicantCharacterReference.save()

            return HttpResponseRedirect(reverse_lazy("applicant_side_job_posts_all"))
        else:
            messages.error(request, "Form Error!")
    

    context = {
        'formApplicantProfile': formApplicantProfile,
        'formApplicantSiblings': formApplicantSiblings,
        'formApplicantEmploymentHistory': formApplicantEmploymentHistory,
        'formApplicantEducationalAttainment': formApplicantEducationalAttainment,
        'formApplicantTrainings': formApplicantTrainings,
        'formApplicantCharacterReference': formApplicantCharacterReference,
    }

    return render(request, template_name, context)

"""
==============================================
********APPLICANT JOB POST PAGE*********
==============================================
"""
# all jobs
class PRIApplicantSideJobPostPageAll(LoginRequiredMixin, TemplateView):
    template_name = 'pri_applicant_side/applicant_job_post_page.html'
    
    def get_context_data(self, *args, **kwargs):
        context = super(PRIApplicantSideJobPostPageAll, self).get_context_data(*args, **kwargs)
        user = User.objects.all().get(username=self.request.user) 

        try:                    
            profile = PRIApplicantProfileInfo.objects.all().filter(user=user)  

            if user.is_staff or user.is_superuser:
                raise Http404()
       
            if not str(self.request.user.username) == user.username:
                return HttpResponseRedirect(reverse_lazy('user_login'))
                #raise Http404()      
        except User.DoesNotExist: 
            return HttpResponseRedirect(reverse_lazy('user_login'))
        
        if profile:
            job_request = PRIApplicantJobRequestInfo.objects.all().filter(applying_applicant__in=profile).order_by('-id')            
            job_vacancies = PRIJobVacancyInfo.objects.exclude(job_vacancy_applied_fk__in=job_request).order_by('-id').distinct()
            test =  PRIJobVacancyInfo.objects.exclude(job_vacancy_applied_fk__in=job_request).filter(job_title__icontains='a').count()
            print(test)
            context.update({
                'user': user,
                'profile': profile, 
                'job_vacancies': job_vacancies, 
                'current_post': 'all',
            })

            return context



# recommendation
class PRIApplicantSideJobPostPageRecommended(LoginRequiredMixin, TemplateView):

    template_name = 'pri_applicant_side/applicant_job_post_page.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PRIApplicantSideJobPostPageRecommended, self).get_context_data(*args, **kwargs)
        #if cookie exp it will redirect to login page don't enclose in try except
        user = User.objects.all().get(username=self.request.user) 

        if user.is_staff or user.is_superuser:
            raise Http404()

        try:                    
            profile = PRIApplicantProfileInfo.objects.all().filter(user=user)  
            employment = PRIApplicantEmploymentHistoryInfo.objects.all().filter(applicant_employment_history__in=profile)
            education = PRIApplicantEducationalAttainmentInfo.objects.all().filter(applicant_educational_attainment__in=profile)
            trainings = PRIApplicantTrainingsInfo.objects.all().filter(applicant_trainings__in=profile)          
            
            if not str(self.request.user.username) == user.username:
                return HttpResponseRedirect(reverse_lazy('user_login'))
                #raise Http404()      
        except User.DoesNotExist: 
            return HttpResponseRedirect(reverse_lazy('user_login'))
            #raise Http404()       

       
        if profile:            
            # Convert Applicant details from model to list
            emp_postion = PRIApplicantEmploymentHistoryInfo.objects.filter(applicant_employment_history__in=profile).values_list('position', flat=True)
            #emp_salary  = PRIApplicantEmploymentHistoryInfo.objects.filter(applicant_employment_history__in=profile).values_list('basic_salary_pay', flat=True) # confidential and decimal format must be in different filtering
            edu_col = PRIApplicantEducationalAttainmentInfo.objects.filter(applicant_educational_attainment__in=profile).values_list('college_course', flat=True)
            edu_voc = PRIApplicantEducationalAttainmentInfo.objects.filter(applicant_educational_attainment__in=profile).values_list('vocational_course', flat=True)
            training_name = PRIApplicantTrainingsInfo.objects.filter(applicant_trainings__in=profile).values_list('title', flat=True)
            all_models_to_filter = list(chain(emp_postion,edu_col,edu_voc,training_name))
            #https://www.geeksforgeeks.org/reduce-in-python/
            #https://docs.python.org/3/library/operator.html
            #http://www.pythonlake.com/python/operators/operatorand_
            #https://stackoverflow.com/questions/4824759/django-query-using-contains-each-value-in-a-list
            #https://stackoverflow.com/questions/16689542/django-count-matches-query-with-contains

            # query_job_qualifications = reduce(operator.or_, (Q(job_qualifications__icontains=x) for x in all_models_to_filter))
            # query_job_responsibilities = reduce(operator.or_, (Q(job_responsibilities=x) for x in all_models_to_filter))
            query_job_qualifications = reduce(operator.or_, (Q(job_qualifications__icontains=x) for x in all_models_to_filter))
            query_job_responsibilities = reduce(operator.or_, (Q(job_responsibilities=x) for x in all_models_to_filter))
            query_job_vacancies = reduce(operator.or_, (Q(Q(job_title__icontains=x) | Q(job_specialization__icontains=x) | Q(job_vacancy_jq_fk__job_qualifications__icontains=x) | Q(job_vacancy_jr_fk__job_responsibilities__icontains=x)) for x in all_models_to_filter))
            

            #job_vacancies = PRIJobVacancyInfo.objects.all().order_by('-id')
            job_request = PRIApplicantJobRequestInfo.objects.all().filter(applying_applicant__in=profile).order_by('-id')            
            job_vacancies = PRIJobVacancyInfo.objects.exclude(job_vacancy_applied_fk__in=job_request).filter(query_job_vacancies).order_by('-id').distinct()
            job_qualifications = PRIJobVacancyJobQualificationsInfo.objects.all().filter(query_job_qualifications).order_by('-id').distinct()
            job_responsibilities = PRIJobVacancyJobResponsibilitiesInfo.objects.all().filter(query_job_responsibilities).order_by('-id').distinct()
            
            # print(query_job_vacancies)
            # print(job_vacancies)
            #return not equals
            #job_request = PRIApplicantJobRequestInfo.objects.all().filter(~Q(applying_applicant__in=profile)).order_by('-id')
            # Excluding the applied job to be listed
            # job_request = PRIApplicantJobRequestInfo.objects.all().filter(applying_applicant__in=profile).order_by('-id')
            # job_vacancies2 = PRIJobVacancyInfo.objects.exclude(job_vacancy_applied_fk__in=job_request)

            # print(job_request)
            # print(job_vacancies2)



            """
            To best explain how this works, a few simple examples might be instructive in 
            helping you understand the code you have. If you are going to continue 
            working with Python code, you will come across list comprehension again, and 
            you may want to use it yourself.

            Note, in the example below, both code segments are equivalent in that they
            create a list of values stored in list myList.

            For instance:

            myList = []
            for i in range(10):
                myList.append(i)

            is equivalent to:

            myList = [i for i in range(10)]

            List comprehensions can be more complex too, so for instance if you had some 
            condition that determined if values should go into a list you could also 
            express this with list comprehension.

            This example only collects even numbered values in the list:

            myList = []
            for i in range(10):
                if i%2 == 0:     # could be written as "if not i%2" more tersely
                myList.append(i)

            and the equivalent list comprehension:

            myList = [i for i in range(10) if i%2 == 0]

            Two final notes:

            1.) You can have "nested" list comrehensions, but they quickly become hard to comprehend :)
            2.) List comprehension will run faster than the equivalent for-loop, and therefore is often 
            a favorite with regular Python programmers who are concerned about efficiency

            Ok, one last example showing that you can also apply functions to the items you are 
            iterating over in the list. This uses float() to convert a list of strings to float values:

            data = ['3', '7.4', '8.2']
            new_data = [float(n) for n in data]

            gives:
            
            new_data
            [3.0, 7.4, 8.2]

            print(reduce(operator.and_, (Q(first_name__contains=x) for x in ['x', 'y', 'z']))) 
            single_liner_loop = [i ** 3 for i in range(0, 1+1, 3)]
            single_liner_loop = [x for x in ['x', 'y', 'z']
            single_liner_loop = [Q(first_name__contains=x) for x in ['x', 'y', 'z']]
            print(single_liner_loop)
            
            """
        
            # for vacancies in job_vacancies:
            #     jq = PRIJobVacancyJobQualificationsInfo.objects.all().filter(job_vacancy_jq=vacancies).order_by('-id').distinct()
            #     jr = PRIJobVacancyJobResponsibilitiesInfo.objects.all().filter(job_vacancy_jr=vacancies).order_by('-id').distinct()
            #
            #     for q in jq:
            #         print(q)
            #     print('---------------')
            #     for r in jr:
            #         print(r)
            #     print(vacancies.job_vacancy_jr_fk)


            """
            How to iterate/filter nested dynamic models in templates
            """

            #for job in job_vacancies:
                #this is the parent iterated from filter queryset
                #accessing the child element by using related name foreign key
                # (.all) is critically needed to call all the queryset from the iterated parent
                # in template use .all only without () but in views include ()
                # for qualifications in job.job_vacancy_jq_fk.all():
                #     print(qualifications)

                # for responisibilites in job.job_vacancy_jr_fk.all():
                #     print(responisibilites)

            context.update({
                'user': user,
                'profile': profile,
                #'profile_user': profile_user,
                'job_vacancies': job_vacancies,
                'job_qualifications': job_qualifications,
                'job_responsibilities': job_responsibilities,
                'current_post': 'recommended',
            })
        return context

@login_required
def PRIApplicantSideJobPostSearchPage(request):
    template_name = 'pri_applicant_side/applicant_job_post_page.html'
    user = User.objects.all().get(username=request.user)

    if user.is_staff or user.is_superuser:
        raise Http404()

    profile_user = get_object_or_404(PRIApplicantProfileInfo, user=user)
    try:                            
        profile = PRIApplicantProfileInfo.objects.all().filter(user=user)
        if not str(request.user.username) == user.username:
            return HttpResponseRedirect(reverse_lazy('user_login'))
            #raise Http404()      
    except User.DoesNotExist:
        return HttpResponseRedirect(reverse_lazy('user_login'))
        #raise Http404()       


    if request.method == 'GET':
        search_term = request.GET.get('search_term')
        #check if only spaces
        if search_term.strip(): 
            if profile:                  
                emp_postion = PRIApplicantEmploymentHistoryInfo.objects.filter(applicant_employment_history__in=profile).values_list('position', flat=True)
                edu_col = PRIApplicantEducationalAttainmentInfo.objects.filter(applicant_educational_attainment__in=profile).values_list('college_course', flat=True)
                edu_voc = PRIApplicantEducationalAttainmentInfo.objects.filter(applicant_educational_attainment__in=profile).values_list('vocational_course', flat=True)
                training_name = PRIApplicantTrainingsInfo.objects.filter(applicant_trainings__in=profile).values_list('title', flat=True)
                all_models_to_filter = list(chain(emp_postion,edu_col,edu_voc,training_name))

                query_job_vacancies = reduce(operator.or_, (Q(Q(job_title__icontains=x) | Q(job_specialization__icontains=x) | Q(job_vacancy_jq_fk__job_qualifications__icontains=x) | Q(job_vacancy_jr_fk__job_responsibilities__icontains=x)) for x in all_models_to_filter))
                #filtering all applied jobs so that it wont be appear
                job_request = PRIApplicantJobRequestInfo.objects.all().filter(applying_applicant__in=profile).order_by('-id')     
                # this is correct but from due issues temporary disabled
                #job_vacancies = PRIJobVacancyInfo.objects.all().exclude(job_vacancy_applied_fk__in=job_request).filter(Q(Q(job_title__icontains=search_term) | Q(job_specialization__icontains=search_term) | Q(client_request__requested_client__client_location__icontains=search_term)) & query_job_vacancies).order_by('-id').distinct()
                job_vacancies = PRIJobVacancyInfo.objects.all().exclude(job_vacancy_applied_fk__in=job_request).filter(Q(Q(job_title__icontains=search_term) | Q(job_specialization__icontains=search_term) | Q(client_request__requested_client__client_location__icontains=search_term))).order_by('-id').distinct()
                context = {
                    'search_term': search_term,
                    'profile': profile,
                    'profile_user': profile_user,
                    'job_vacancies': job_vacancies, 
                    'current_post': 'all',
                }

                return render(request, template_name, context=context)
            else:
                return HttpResponseRedirect(reverse_lazy('applicant_side_job_posts_all'))
        else:
            return HttpResponseRedirect(reverse_lazy('applicant_side_job_posts_all'))
    else:
        return redirect('applicant_side_job_posts_all')

@login_required
def PRIApplicantSideFilterCategoryPage(request):
    template_name = 'pri_applicant_side/applicant_job_post_page.html'
    user = User.objects.all().get(username=request.user)

    if user.is_staff or user.is_superuser:
        raise Http404()

    profile_user = get_object_or_404(PRIApplicantProfileInfo, user=user)
    try:                            
        profile = PRIApplicantProfileInfo.objects.all().filter(user=user)
        if not str(request.user.username) == user.username:
            return HttpResponseRedirect(reverse_lazy('user_login'))
            #raise Http404()      
    except User.DoesNotExist:
        return HttpResponseRedirect(reverse_lazy('user_login'))
        #raise Http404()       

    if request.method == "GET":        
        experience = request.GET.get('experience')
        skills = request.GET.get('skills')
        education = request.GET.get('education')  

        profile_match = []

        exp = PRIApplicantEmploymentHistoryInfo.objects.all().filter(applicant_employment_history__in=profile)
        edu = PRIApplicantEducationalAttainmentInfo.objects.all().filter(applicant_educational_attainment__in=profile)

        if experience:
            for e in exp:  
                profile_match.append(e.company)
                profile_match.append(e.position)
        
        if education:
            for e in edu:
                profile_match.append(e.college)
                profile_match.append(e.college_course)
                profile_match.append(e.vocational)
                profile_match.append(e.vocational_course)
                profile_match.append(e.elementary)
                profile_match.append(e.highschool) 
        if skills:
            for e in edu: 
                profile_match.append(e.special_skills)  

        
        job_request = PRIApplicantJobRequestInfo.objects.all().filter(applying_applicant__in=profile).order_by('-id')            
        job_vacancies = PRIJobVacancyInfo.objects.all().exclude(job_vacancy_applied_fk__in=job_request).order_by('-id').distinct()

        if experience or education or skills:
            query_job_vacancies = reduce(operator.or_, (Q(Q(job_title__icontains=x) | Q(job_specialization__icontains=x) | Q(job_vacancy_jq_fk__job_qualifications__icontains=x) | Q(job_vacancy_jr_fk__job_responsibilities__icontains=x)) for x in profile_match))
            job_vacancies = PRIJobVacancyInfo.objects.exclude(job_vacancy_applied_fk__in=job_request).filter(query_job_vacancies).order_by('-id').distinct()
            print(query_job_vacancies)       
        
        context = {
            'profile': profile,
            'profile_user': profile_user,
            'experience': experience,
            'skills': skills,
            'education': education,
            'job_vacancies': job_vacancies,
        }
        return render(request, template_name, context=context)
    else:
        return redirect('applicant_side_job_posts_all')
"""
-----------------------------
    APPLICANT JOB REQUEST
-----------------------------
"""

def pri_applicant_job_request(request, jv_id, a_id):    
    data = dict()
    template_name = "pri_applicant_side/applicant_job_request.html"
    job_vacancy = get_object_or_404(PRIJobVacancyInfo, id=jv_id)
    applicant = get_object_or_404(PRIApplicantProfileInfo, id=a_id)
    if request.is_ajax():
        if request.method == 'POST':
            # add record to model without using form
            PRIApplicantJobRequestInfo.objects.create(job_vacancy_applied=job_vacancy, applying_applicant=applicant)
            data['form_is_valid'] = True  
        elif request.method == 'GET':
            context = {
                'job_vacancy': job_vacancy,
                'applicant': applicant,
            } 
            data['html_form'] = render_to_string(template_name, context, request) 
        return JsonResponse(data)
    else:
        raise Http404()

"""
---------------------------------------
    Applicant Job Request Status
---------------------------------------
"""

def pri_applicant_job_request_ongoing_status(request, id):
    data = dict()
    template_name = "pri_applicant_side/applicant_job_request_ongoing_status.html"
    #profile = PRIApplicantProfileInfo.objects.all().filter(id=id).order_by('-id').distinct()
    profile = get_object_or_404(PRIApplicantProfileInfo, id=id)
    if request.is_ajax():
        if request.method == 'GET':
            job_request = PRIApplicantJobRequestInfo.objects.all().filter(applying_applicant=profile).order_by('-id').distinct()             
            context = {
                'job_request': job_request,
                'profile': profile,
            }
            data['html_form'] = render_to_string(template_name, context, request)
            print(job_request)
            return JsonResponse(data)            

        else:
            return HttpResponseRedirect(reverse_lazy('applicant_side_job_posts_all'))
    else:
        raise Http404()


def pri_applicant_job_request_cancel(request):
    data = dict()

    if request.is_ajax():
        if request.method == "POST":

            json_data = json.loads(request.body)

            print(json.dumps(json_data, indent=4, sort_keys=True))
            for id in json_data['ids']:
                record = get_object_or_404(PRIApplicantJobRequestInfo, id=id)
                record.delete()
                print('->', id)
            data['valid'] = True
            return JsonResponse(data)
    else:
        raise Http404()

def pri_applicant_job_hired(request, id):
    data = dict()
    template_name = "pri_applicant_side/applicant_job_hired.html"
    profile = get_object_or_404(PRIApplicantProfileInfo, id=id)
    company = PRIApplicantJobHiredInfo.objects.all().filter(hired_applicant=profile).order_by('-id').distinct()

    if request.is_ajax():
        if request.method == 'GET':

            context = {
                'profile': profile,
                'company': company,
            }

            data['html_form'] = render_to_string(template_name, context, request)
            return JsonResponse(data)  
    else:
        raise Http404()


"""
---------------------------------------
******Applicant Examninations**********
---------------------------------------
"""

class PRIApplicantExamListPage(LoginRequiredMixin, TemplateView):
    template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_list.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PRIApplicantExamListPage, self).get_context_data(*args, **kwargs)
        
        key = decrypt_key(self.kwargs.get('key'))
        profile = get_object_or_404(PRIApplicantProfileInfo, id=key)
        job_vacancy = get_object_or_404(PRIApplicantJobRequestInfo, id=self.kwargs.get('id'))

        t1pe_score = PRIApplicantExamScoreT1PEInfo.objects.all().filter(applicant_exam_score_t1pe=profile, job_request_exam_score_t1pe=job_vacancy)
        t2se_score = PRIApplicantExamScoreT2SEInfo.objects.all().filter(applicant_exam_score_t2se=profile, job_request_exam_score_t2se=job_vacancy)
        t3e_score = PRIApplicantExamScoreT3EInfo.objects.all().filter(applicant_exam_score_t3e=profile, job_request_exam_score_t3e=job_vacancy)
        t3ptsct_score = PRIApplicantExamScoreT3PTSCTInfo.objects.all().filter(applicant_exam_score_t3ptsct=profile, job_request_exam_score_t3ptsct=job_vacancy)
        t4ar_score = PRIApplicantExamScoreT4ARInfo.objects.all().filter(applicant_exam_score_t4ar=profile, job_request_exam_score_t4ar=job_vacancy)
        cca_score = PRIApplicantExamScoreCCAInfo.objects.all().filter(applicant_exam_score_cca=profile, job_request_exam_score_cca=job_vacancy)
        arp_score = PRIApplicantExamScoreARPInfo.objects.all().filter(applicant_exam_score_arp=profile, job_request_exam_score_arp=job_vacancy)
        scct_score = PRIApplicantExamScoreSCCTInfo.objects.all().filter(applicant_exam_score_scct=profile, job_request_exam_score_scct=job_vacancy)
        
        final_score = 0
        final_over = 0

        for t1pe in t1pe_score: 
            final_score += t1pe.score_t1pe
            final_over += t1pe.over_t1pe

        for t2se in t2se_score: 
            final_score += t2se.score_t2se
            final_over += t2se.over_t2se
        
        for t3e in t3e_score: 
            final_score += t3e.score_t3e
            final_over += t3e.over_t3e
        
        for t3ptsct in t3ptsct_score: 
            final_score += t3ptsct.score_t3ptsct
            final_over += t3ptsct.over_t3ptsct
        
        for t4ar in t4ar_score: 
            final_score += t4ar.score_t4ar
            final_over += t4ar.over_t4ar
        
        for cca in cca_score: 
            final_score += cca.score_cca
            final_over += cca.over_cca
        
        for arp in arp_score: 
            final_score += arp.score_arp
            final_over += arp.over_arp
        
        for scct in scct_score: 
            final_score += scct.score_scct
            final_over += scct.over_scct
        
 
        
        if not job_vacancy.take_exam:
            raise Http404()

        context.update({
            'profile': profile,
            'job_vacancy': job_vacancy,
            't1pe_score': t1pe_score,
            't2se_score': t2se_score,
            't3e_score': t3e_score,
            't3ptsct_score': t3ptsct_score,
            't4ar_score': t4ar_score,
            'cca_score': cca_score,
            'arp_score': arp_score,
            'scct_score': scct_score,
            'final_score': final_score,
            'final_over': final_over,
            'final_passing_score': (final_over/2)
        })

        return context

 

# class PRIApplicantTest1PreEmployment(LoginRequiredMixin, TemplateView):

#     template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_test1.html'

def calculate_exam_score(score, over): 
    
    exam_score = score/over
    ratings = 'Poor'
    status = 'Failed'

    if exam_score > 0 and exam_score < 0.25:
        ratings = 'Poor'
        status = 'Failed'
    elif exam_score > 0.25 and exam_score < 0.50:
        ratings = 'Below Average'
        status = 'Failed'
    elif exam_score > 0.50 and exam_score < 0.75:
        ratings = 'Average'
        status = 'Passed'
    elif exam_score > 0.75 and exam_score <= 1.0:
        ratings = 'Outstanding'
        status = 'Passed'
    print(ratings, status)
    return ratings, status

     

@login_required
def pri_applicant_test1_pre_employment(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_test1.html'
    key = decrypt_key(key)
    profile = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIApplicantJobRequestInfo, id=id)

    if not job_vacancy.take_exam:
        raise Http404() 

    test = PRIApplicantExamScoreT1PEInfo.objects.all().filter(applicant_exam_score_t1pe=profile, job_request_exam_score_t1pe=job_vacancy)
    if test:
        for field in test:
            if not field.allow_retake_t1pe:
                raise Http404()

    if request.is_ajax():
        if request.method == 'POST':
            json_data = json.loads(request.body)
            
            print(json.dumps(json_data, indent=4, sort_keys=True))

            score = (json_data['score'] * 2)
            over = (json_data['over'] * 2)

            ratings, status = calculate_exam_score(score, over) 
            
            t1pe, created = PRIApplicantExamScoreT1PEInfo.objects.update_or_create(applicant_exam_score_t1pe=profile, job_request_exam_score_t1pe=job_vacancy, 
            defaults={'score_t1pe': score, 'over_t1pe' : over, 'status_t1pe': status, 'ratings_t1pe': ratings, 'allow_retake_t1pe': False})

            if created:
                print('Already Exists!')
            else:
                print('New Record')

            data['urlAttr'] = reverse_lazy('pri:applicant_exam_list', kwargs={
                'key': profile.secure_key_id,
                'id': job_vacancy.id
            })

            # print(HttpResponseRedirect(reverse_lazy('pri:applicant_exam_list', kwargs={
            #     'key': profile.secure_key_id,
            #     'id': job_vacancy.id
            # })))

            return JsonResponse(data)
            
            # return HttpResponseRedirect(reverse_lazy('pri:applicant_exam_list', kwargs={
            #     'key': profile.secure_key_id,
            #     'id': job_vacancy.id
            # }))
    else: 
        context = {
            'profile': profile,
            'job_vacancy': job_vacancy,
        }

        return render(request, template_name, context)
 

@login_required
def pri_applicant_test2_simple_english(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_test2.html'
    key = decrypt_key(key)
    profile = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIApplicantJobRequestInfo, id=id)

    if not job_vacancy.take_exam:
        raise Http404() 

    test = PRIApplicantExamScoreT2SEInfo.objects.all().filter(applicant_exam_score_t2se=profile, job_request_exam_score_t2se=job_vacancy)
    if test:
        for field in test:
            if not field.allow_retake_t2se:
                raise Http404()

    if request.is_ajax():
        if request.method == 'POST':
            json_data = json.loads(request.body)
            print(json.dumps(json_data, indent=4, sort_keys=True))

            score = (json_data['score'] * 2)
            over = (json_data['over'] * 2)

            ratings, status = calculate_exam_score(score, over) 
 

            t2se, created = PRIApplicantExamScoreT2SEInfo.objects.update_or_create(applicant_exam_score_t2se=profile, job_request_exam_score_t2se=job_vacancy, 
            defaults={'score_t2se': score, 'over_t2se' : over, 'status_t2se': status, 'ratings_t2se': ratings, 'allow_retake_t2se': False})

            if created:
                print('Already Exists!')
            else:
                print('New Record')

            data['urlAttr'] = reverse_lazy('pri:applicant_exam_list', kwargs={
                'key': profile.secure_key_id,
                'id': job_vacancy.id
            }) 

            return JsonResponse(data)
    else: 
        context = {
            'profile': profile,
            'job_vacancy': job_vacancy,
        }

        return render(request, template_name, context)

@login_required
def pri_applicant_test3_essay(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_test3.html'
    key = decrypt_key(key)
    profile = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIApplicantJobRequestInfo, id=id)

    if not job_vacancy.take_exam:
        raise Http404() 

    test = PRIApplicantExamScoreT3EInfo.objects.all().filter(applicant_exam_score_t3e=profile, job_request_exam_score_t3e=job_vacancy)
    if test:
        for field in test:
            if not field.allow_retake_t3e:
                raise Http404()
 
    if request.method == 'POST':
        essayForm = PRIApplicantTest3EssayForm(request.POST or None)

        if essayForm.is_valid():
            # https://docs.djangoproject.com/en/dev/ref/models/querysets/#get-or-create
            #https://docs.djangoproject.com/en/dev/ref/models/querysets/#update-or-create
            # print(essayForm.cleaned_data)
            # print(essayForm.cleaned_data['question1'])
            essay, created = PRIApplicantTest3EssayInfo.objects.update_or_create(applicant_exam=profile, job_request_exam=job_vacancy, defaults={'question1':essayForm.cleaned_data['question1'], 'question2' : essayForm.cleaned_data['question2'], 'question3' : essayForm.cleaned_data['question3']})
            # essay, created = PRIApplicantTest3EssayInfo.objects.update_or_create(applicant_exam=profile, job_request_exam=job_vacancy, question1=essayForm.cleaned_data['question1'], question2 =essayForm.cleaned_data['question2'], question3 = essayForm.cleaned_data['question3'])
            # essay, created = PRIApplicantTest3EssayInfo.objects.get_or_create(**essayForm.cleaned_data)

            t3e, created_t3e = PRIApplicantExamScoreT3EInfo.objects.update_or_create(applicant_exam_score_t3e=profile, job_request_exam_score_t3e=job_vacancy, 
            defaults={'score_t3e': 0, 'over_t3e' : 30, 'status_t3e': 'On progress', 'ratings_t3e': 'On progress', 'allow_retake_t3e': False})
            
            if created:
                print("Already Exists!")
            else:
                print("New Record!")

            # instance = essayForm.save(commit=False)
            # instance.applicant_exam = profile
            # instance.job_request_exam = job_vacancy
            # instance.save()
            return HttpResponseRedirect(reverse_lazy('pri:applicant_exam_list', kwargs={
            'key': profile.secure_key_id,
            'id': job_vacancy.id
            }))       
    elif request.method == 'GET': 
        essayForm = PRIApplicantTest3EssayForm(request.GET or None)

    context = {
        'profile': profile,
        'job_vacancy': job_vacancy,
        'essayForm': essayForm,
    }

    return render(request, template_name, context)


@login_required
def pri_applicant_test3_sct(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_test3_sct.html'
    key = decrypt_key(key)
    profile = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIApplicantJobRequestInfo, id=id)

    if not job_vacancy.take_exam:
        raise Http404() 

    test = PRIApplicantExamScoreT3PTSCTInfo.objects.all().filter(applicant_exam_score_t3ptsct=profile, job_request_exam_score_t3ptsct=job_vacancy)
    if test:
        for field in test:
            if not field.allow_retake_t3ptsct:
                raise Http404()
 
    if request.method == 'POST':
        sctForm = PRIApplicantTest3SCTForm(request.POST or None)      
        if sctForm.is_valid():  
            print(sctForm.cleaned_data)
            sct, created = PRIApplicantTest3SCTInfo.objects.update_or_create(applicant_exam_sct=profile, job_request_exam_sct=job_vacancy, defaults=sctForm.cleaned_data)
            
            t3ptsct, created_t3ptsct = PRIApplicantExamScoreT3PTSCTInfo.objects.update_or_create(applicant_exam_score_t3ptsct=profile, job_request_exam_score_t3ptsct=job_vacancy, 
            defaults={'score_t3ptsct': 0, 'over_t3ptsct' : 15, 'status_t3ptsct': 'On progress', 'ratings_t3ptsct': 'On progress', 'allow_retake_t3ptsct': False})
            
            if created:
                print("Already Exists!")
            else:
                print("New Record!") 
            return HttpResponseRedirect(reverse_lazy('pri:applicant_exam_list', kwargs={
            'key': profile.secure_key_id,
            'id': job_vacancy.id
            }))       
    elif request.method == 'GET': 
        sctForm = PRIApplicantTest3SCTForm(request.GET or None)

    context = {
        'profile': profile,
        'job_vacancy': job_vacancy,
        'sctForm': sctForm,
    }

    return render(request, template_name, context)


@login_required
def pri_applicant_test4_abstract_reasoning(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_test4.html'
    key = decrypt_key(key)
    profile = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIApplicantJobRequestInfo, id=id)

    if not job_vacancy.take_exam:
        raise Http404() 

    test = PRIApplicantExamScoreT4ARInfo.objects.all().filter(applicant_exam_score_t4ar=profile, job_request_exam_score_t4ar=job_vacancy)
    if test:
        for field in test:
            if not field.allow_retake_t4ar:
                raise Http404()

    if request.is_ajax():
        if request.method == 'POST': 
            json_data = json.loads(request.body)
            print(json.dumps(json_data, indent=4, sort_keys=True))

            score = json_data['score'] 
            over = json_data['over']

            ratings, status = calculate_exam_score(score, over)  
            
            t4ar, createdss = PRIApplicantExamScoreT4ARInfo.objects.update_or_create(applicant_exam_score_t4ar=profile, job_request_exam_score_t4ar=job_vacancy, 
            defaults={'score_t4ar': score, 'over_t4ar' : over, 'status_t4ar': status, 'ratings_t4ar': ratings, 'allow_retake_t4ar': False})

            if createdss:
                print("Already Exists!")
            else:
                print("New Record!") 

            data['urlAttr'] = reverse_lazy('pri:applicant_exam_list', kwargs={
                'key': profile.secure_key_id,
                'id': job_vacancy.id
            }) 
            return JsonResponse(data)
    else: 
        context = {
            'profile': profile,
            'job_vacancy': job_vacancy,
        }

        return render(request, template_name, context)
    

@login_required
def pri_applicant_exam_ssct(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_ssct.html'
    key = decrypt_key(key)
    profile = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIApplicantJobRequestInfo, id=id)

    if not job_vacancy.take_exam:
        raise Http404() 

    if job_vacancy.job_vacancy_applied.job_category != 'Janitor':
        raise Http404()

    test = PRIApplicantExamScoreSCCTInfo.objects.all().filter(applicant_exam_score_scct=profile, job_request_exam_score_scct=job_vacancy)
    if test:
        for field in test:
            if not field.allow_retake_scct:
                raise Http404()
 
    if request.method == 'POST':
        scctForm = PRIApplicantTestSSCTForm(request.POST or None)
        if scctForm.is_valid():  
            print(scctForm.cleaned_data)
            scct, created = PRIApplicantTestSSCTInfo.objects.update_or_create(applicant_exam_ssct=profile, job_request_exam_ssct=job_vacancy, defaults=scctForm.cleaned_data)
            
            sccts, created_scct = PRIApplicantExamScoreSCCTInfo.objects.update_or_create(applicant_exam_score_scct=profile, job_request_exam_score_scct=job_vacancy, 
            defaults={'score_scct': 0, 'over_scct' : 60, 'status_scct': 'On progress', 'ratings_scct': 'On progress', 'allow_retake_scct': False})
            
            if created:
                print("Already Exists!")
            else:
                print("New Record!") 
            return HttpResponseRedirect(reverse_lazy('pri:applicant_exam_list', kwargs={
            'key': profile.secure_key_id,
            'id': job_vacancy.id
            }))    
    elif request.method == 'GET': 
        scctForm = PRIApplicantTestSSCTForm(request.GET or None)

    context = {
        'profile': profile,
        'job_vacancy': job_vacancy,
        'scctForm': scctForm,
    }

    return render(request, template_name, context)


@login_required
def pri_applicant_exam_ccat(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_ccat.html'
    key = decrypt_key(key)
    profile = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIApplicantJobRequestInfo, id=id)

    if not job_vacancy.take_exam:
        raise Http404()  

    test = PRIApplicantExamScoreCCAInfo.objects.all().filter(applicant_exam_score_cca=profile, job_request_exam_score_cca=job_vacancy)
    if test:
        for field in test:
            if not field.allow_retake_cca:
                raise Http404()

 
    if request.method == 'POST':
        ccatForm = PRIApplicantTestCCATForm(request.POST or None)
        if ccatForm.is_valid():   
            ccat, created = PRIApplicantTestCCATInfo.objects.update_or_create(applicant_exam_ccat=profile, job_request_exam_ccat=job_vacancy, defaults=ccatForm.cleaned_data)

            sccts, created_ccats = PRIApplicantExamScoreCCAInfo.objects.update_or_create(applicant_exam_score_cca=profile, job_request_exam_score_cca=job_vacancy, 
            defaults={'score_cca': 0, 'over_cca' : 60, 'status_cca': 'On progress', 'ratings_cca': 'On progress', 'allow_retake_cca': False})
            
            
            if created:
                print("Already Exists!")
            else:
                print("New Record!") 
            return HttpResponseRedirect(reverse_lazy('pri:applicant_exam_list', kwargs={
            'key': profile.secure_key_id,
            'id': job_vacancy.id
            }))    
    elif request.method == 'GET': 
        ccatForm = PRIApplicantTestCCATForm(request.GET or None)

    context = {
        'profile': profile,
        'job_vacancy': job_vacancy,
        'ccatForm': ccatForm,
    }

    return render(request, template_name, context)


@login_required
def pri_applicant_exam_arp(request, key, id):
    data = dict()
    template_name = 'pri_admin/pri_admin_applicant_exams/pri_admin_applicant_exam_arp.html'
    key = decrypt_key(key)
    profile = get_object_or_404(PRIApplicantProfileInfo, id=key)
    job_vacancy = get_object_or_404(PRIApplicantJobRequestInfo, id=id)

    if not job_vacancy.take_exam:
        raise Http404()  
        
    test = PRIApplicantExamScoreARPInfo.objects.all().filter(applicant_exam_score_arp=profile, job_request_exam_score_arp=job_vacancy)
    if test:
        for field in test:
            if not field.allow_retake_arp:
                raise Http404()

    if request.method == 'POST':
        arpForm = PRIApplicantTestARPForm(request.POST or None)
        if arpForm.is_valid():   
            arp, created = PRIApplicantTestARPInfo.objects.update_or_create(applicant_exam_arp=profile, job_request_exam_arp=job_vacancy, defaults=arpForm.cleaned_data)

            arps, created_arps = PRIApplicantExamScoreARPInfo.objects.update_or_create(applicant_exam_score_arp=profile, job_request_exam_score_arp=job_vacancy, 
            defaults={'score_arp': 0, 'over_arp' : 65, 'status_arp': 'On progress', 'ratings_arp': 'On progress', 'allow_retake_arp': False})
            
            if created:
                print("Already Exists!")
            else:
                print("New Record!") 
            return HttpResponseRedirect(reverse_lazy('pri:applicant_exam_list', kwargs={
            'key': profile.secure_key_id,
            'id': job_vacancy.id
            }))    
    elif request.method == 'GET': 
        arpForm = PRIApplicantTestARPForm(request.GET or None)

    context = {
        'profile': profile,
        'job_vacancy': job_vacancy,
        'arpForm': arpForm,
    }

    return render(request, template_name, context)



# Client side

class PRIClientHomePage(LoginRequiredMixin, TemplateView):
    template_name = 'pri_client_side/client_home_page.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PRIClientHomePage, self).get_context_data(*args, **kwargs)
        user = User.objects.all().get(username=self.request.user) 
        
        if not user.is_staff or user.is_superuser:
            raise Http404()
        
        try:                    
            profile = PRIUserInfo.objects.all().filter(user=user)    
            client = PRIClientsInfo.objects.all().filter(client_user__in=profile)
            requests = PRIRequestInfo.objects.all().filter(requested_client__in=client).order_by('-id').distinct()
            
           
            if not str(self.request.user.username) == user.username:
                return HttpResponseRedirect(reverse_lazy('user_login'))
            
        except PRIUserInfo.DoesNotExist: 
            return HttpResponseRedirect(reverse_lazy('user_login'))

        context.update({
            'user': user,
            'profile': profile,
            'client': client,
            'requests': requests,
        })

        return context

def pri_client_side_create_requests(request):
    data = dict()
    template_name = 'pri_client_side/client_create_request.html'
    user = User.objects.all().get(username=request.user) 
    profile = PRIUserInfo.objects.all().filter(user=user)    
    client = PRIClientsInfo.objects.all().get(client_user__in=profile)

    if request.is_ajax():
        if request.method == 'POST':
            form = PRIClientRequestForm(request.POST)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.status = 'Started'
                instance.requested_client = client
                form.save() 

                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False 

        elif request.method == 'GET':
            form = PRIClientRequestForm() 
        context = {
            'form': form,
        } 
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

def pri_client_side_edit_requests(request, id):
    data = dict()
    template_name = 'pri_client_side/client_edit_request.html' 
    job_request = get_object_or_404(PRIRequestInfo, id=id)

    if request.is_ajax():
        if request.method == 'POST':
            form = PRIClientRequestForm(request.POST or None, instance=job_request)
            if form.is_valid(): 
                instance = form.save(commit=False)
                instance.seen_by_admin = False
                instance.save() 
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False 

        elif request.method == 'GET':
            form = PRIClientRequestForm(request.GET or None, instance=job_request) 
        context = {
            'form': form,
        } 
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


def pri_client_side_delete_requests(request, id):
    data = dict()
    template_name = 'pri_client_side/client_delete_request.html' 
    job_request = get_object_or_404(PRIRequestInfo, id=id)

    if request.is_ajax():
        if request.method == 'POST':
            job_request.delete()
            data['form_is_valid'] = True 

        elif request.method == 'GET': 
            context = {
                'job_request': job_request,
            } 
            data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


class PRIClientHiredApplicantPage(LoginRequiredMixin, TemplateView):
    template_name = 'pri_client_side/client_hired_applicants_page.html'

    def get_context_data(self, *args, **kwargs):
        context = super(PRIClientHiredApplicantPage, self).get_context_data(*args, **kwargs)
        user = User.objects.all().get(username=self.request.user) 

        
        if not user.is_staff or user.is_superuser:
            raise Http404()
        
        try:                    
            profile = PRIUserInfo.objects.all().filter(user=user)    
            client = PRIClientsInfo.objects.all().filter(client_user__in=profile)
            hired_applicants = PRIApplicantJobHiredInfo.objects.all().filter(company_name__in=client).order_by('-id').distinct()
            
            print(hired_applicants)
            if not str(self.request.user.username) == user.username:
                return HttpResponseRedirect(reverse_lazy('user_login'))
            
        except PRIUserInfo.DoesNotExist: 
            return HttpResponseRedirect(reverse_lazy('user_login'))

        context.update({
            'user': user,
            'profile': profile,
            'client': client,
            'hired_applicants': hired_applicants,
        })

        return context


def pri_client_delete_applicants(request, id):
    data = dict()
    template_name = 'pri_client_side/client_fire_applicant.html'
    applicant = get_object_or_404(PRIApplicantJobHiredInfo, id=id)

    if request.is_ajax():
        if request.method == 'POST':
            applicant.delete()
            data['form_is_valid'] = True 
        elif request.method == 'GET':
            context = { 
                'applicant': applicant,
            }
            data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


# reports

def export_report_xlsx_applicants(request):
    # https://openpyxl.readthedocs.io/en/stable/tutorial.html
    # There is no need to create a file on the filesystem to get started with openpyxl. Just import the Workbook class and start work:

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename=applicants-{date}.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

    workbook = Workbook()

    # A workbook is always created with at least one worksheet. You can get it by using the Workbook.active property:

    ws = workbook.active

    # This is set to 0 by default. Unless you modify its value, you will always get the first worksheet by using this method.

    ws = workbook.create_sheet("Applicants")

    # Sheets are given a name automatically when they are created. They are numbered in sequence (Sheet, Sheet1, Sheet2, …). You can change this name at any time with the Worksheet.title property:

    ws.title = "Applicants"

    # The background color of the tab holding this title is white by default. You can change this providing an RRGGBB color code to the Worksheet.sheet_properties.tabColor attribute:

    ws.sheet_properties.tabColor = "1072BA"

    
    thin = Side(border_style="thin", color="007BFF")
    thick = Side(border_style="thick", color="1E1E1E")
    double = Side(border_style="double", color="ff0000")
    # Title
    ws.merge_cells('A1:D1') 
    title_cell = ws['A1'] 
    title_cell.value = "Applicant Reports" 

    title_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    title_cell.fill = PatternFill("solid", fgColor="1E1E1E") 
    title_cell.font = Font(color=colors.WHITE, bold=True, size=20)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust the size
    ws.row_dimensions[1].height = 30 
    # ws.column_dimensions['A'].width = 40
    
 
    columns = ["Username","Name","Mobile No", "Email"]
    alphabetic_chars = string.ascii_uppercase
    # must always start at 1
    for col in range(len(columns)):         
        ws.cell(row=2, column=col+1, value=columns[col])
        ws.row_dimensions[2].height = 30 
        ws.column_dimensions[alphabetic_chars[col]].width = 40

    # Now we know how to get a worksheet, we can start modifying cells content. Cells can be accessed directly as keys of the worksheet:

    header_columns = [ws['A2'],ws['B2'],ws['C2'],ws['D2']]
    ft = Font(color=colors.BLACK, bold=True, size=12)

    """
    https://tutorialspoint.dev/language/python/python-adjusting-rows-and-columns-of-an-excel-file-using-openpyxl-module
    sheet = wb.active 
    # set the height of the row 
    sheet.row_dimensions[1].height = 70
    
    # set the width of the column 
    sheet.column_dimensions['B'].width = 20
    """

    # If you want to change the color of a Font, you need to reassign it::

    for element in header_columns:
        element.font = ft  # the change only affects A1
        element.border = Border(top=thick, left=thick, right=thick, bottom=thick)
        element.fill = PatternFill("solid", fgColor="FFE300")
        element.alignment = Alignment(horizontal="center", vertical="center")

    # ws.merge_cells('B2:F4')

    # top_left_cell = ws['B2']
    # top_left_cell.value = "My Cell" 

    # top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
    # top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
    # top_left_cell.fill = GradientFill(stop=("000000", "FFFFFF"))
    # top_left_cell.font = Font(b=True, color="FF0000")
    # top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

 
    applicants = PRIApplicantProfileInfo.objects.all().order_by('-id').distinct()
    index = 2
    for applicant in applicants:
        index += 1
        ws.cell(row=index, column=1, value=applicant.user.username).border = thin_border
        ws.cell(row=index, column=2, value=str(applicant.lname +", " + applicant.fname + " " + applicant.mname)).border = thin_border
        ws.cell(row=index, column=3, value=applicant.mobile_no).border = thin_border
        ws.cell(row=index, column=4, value=applicant.user.email).border = thin_border

    
    
    workbook.save(response)

    return response

def export_report_xlsx_applying_applicants(reqeuest):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename=applying-applicants-{date}.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

    workbook = Workbook()

    ws = workbook.active

    ws = workbook.create_sheet("Appying Applicants")

    ws.title = "Appying Applicants"

    ws.sheet_properties.tabColor = "1072BA"
    
    thin = Side(border_style="thin", color="007BFF")
    thick = Side(border_style="thick", color="1E1E1E")
    double = Side(border_style="double", color="ff0000") 

    # Adjust the size
    ws.row_dimensions[1].height = 30  
    
    columns = ["ID","Client","Title","Category","Specialization","Min. Exp","Salary", "Deadline","Name","Mobile","Email","Status","Date Applied","Interview Date"]
    alphabetic_chars = string.ascii_uppercase 

    # Title
    merge_cell = 'A1:{}1'.format(alphabetic_chars[len(columns)-1]) 

    ws.merge_cells(merge_cell) 
    title_cell = ws['A1'] 
    title_cell.value = "Applicant Reports" 

    title_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    title_cell.fill = PatternFill("solid", fgColor="1E1E1E") 
    title_cell.font = Font(color=colors.WHITE, bold=True, size=20)
    title_cell.alignment = Alignment(horizontal="center", vertical="center") 

    # must always start at 1
    for col in range(len(columns)):         
        ws.cell(row=2, column=col+1, value=columns[col])
        ws.row_dimensions[2].height = 30 
        ws.column_dimensions[alphabetic_chars[col]].width = 40

    header_columns = [ws.cell(row=2,column=x) for x in range(1,(len(columns)+1))]
    ft = Font(color=colors.BLACK, bold=True, size=12)
 
    for element in header_columns:
        element.font = ft  # the change only affects A1
        element.border = Border(top=thick, left=thick, right=thick, bottom=thick)
        element.fill = PatternFill("solid", fgColor="FFE300")
        element.alignment = Alignment(horizontal="center", vertical="center")

    # Database Query
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    # query = PRIJobVacancyInfo.objects.all().order_by('-id')
    # job_requests = PRIApplicantJobRequestInfo.objects.all().order_by('-id')

    # index = 2
    # for job_request in job_requests:
    #     index += 1
    #     ws.cell(row=index, column=1, value=job_request.job_vacancy_applied.id).border = thin_border
    #     ws.cell(row=index, column=2, value=job_request.job_vacancy_applied.client_request.requested_client.client_company_name).border = thin_border
    #     ws.cell(row=index, column=3, value=job_request.job_vacancy_applied.job_title).border = thin_border
    #     ws.cell(row=index, column=4, value=job_request.job_vacancy_applied.job_category).border = thin_border
    #     ws.cell(row=index, column=5, value=job_request.job_vacancy_applied.job_specialization).border = thin_border
    #     ws.cell(row=index, column=6, value=job_request.job_vacancy_applied.job_minimum_experience).border = thin_border
    #     ws.cell(row=index, column=7, value=job_request.job_vacancy_applied.job_salary).border = thin_border
    #     ws.cell(row=index, column=8, value=job_request.job_vacancy_applied.job_deadline).border = thin_border
        
    job_vacancy = PRIJobVacancyInfo.objects.all().order_by('-id')

    index = 2
    for vacancy in job_vacancy:        
        index += 1        
        ws.cell(row=index, column=1, value=vacancy.id).border = thin_border
        ws.cell(row=index, column=2, value=vacancy.client_request.requested_client.client_company_name).border = thin_border
        ws.cell(row=index, column=3, value=vacancy.job_title).border = thin_border
        ws.cell(row=index, column=4, value=vacancy.job_category).border = thin_border
        ws.cell(row=index, column=5, value=vacancy.job_specialization).border = thin_border
        ws.cell(row=index, column=6, value=vacancy.job_minimum_experience).border = thin_border
        ws.cell(row=index, column=7, value=vacancy.job_salary).border = thin_border
        ws.cell(row=index, column=8, value=vacancy.job_deadline).border = thin_border
        job_requests = PRIApplicantJobRequestInfo.objects.all().filter(job_vacancy_applied=vacancy).order_by('-id')
        for job_request in job_requests: 
            index += 1
            ws.cell(row=index, column=9, value=str(job_request.applying_applicant.lname + ", " + job_request.applying_applicant.fname + " " + job_request.applying_applicant.mname)).border = thin_border
            ws.cell(row=index, column=10, value=job_request.applying_applicant.mobile_no).border = thin_border
            ws.cell(row=index, column=11, value=job_request.applying_applicant.user.email).border = thin_border
            ws.cell(row=index, column=12, value="On Going").border = thin_border
            ws.cell(row=index, column=13, value=job_request.date_applied).border = thin_border
            ws.cell(row=index, column=14, value=job_request.interview_date).border = thin_border
            

    workbook.save(response)

    return response
   
def export_report_xlsx_job_vacants(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename=job-vacants-{date}.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

    workbook = Workbook()

    ws = workbook.active
    
    ws = workbook.create_sheet("Job Vacants")

    ws.title = "Job Vacants"

    ws.sheet_properties.tabColor = "1072BA"
    
    thin = Side(border_style="thin", color="007BFF")
    thick = Side(border_style="thick", color="1E1E1E")
    double = Side(border_style="double", color="ff0000") 

    # Adjust the size
    ws.row_dimensions[1].height = 30  
    
    columns = ["ID","Client","Title","Category","Specialization","Min. Exp","Salary", "Deadline"]
    alphabetic_chars = string.ascii_uppercase 
        
    # Title
    merge_cell = 'A1:{}1'.format(alphabetic_chars[len(columns)-1]) 

    ws.merge_cells(merge_cell) 
    title_cell = ws['A1'] 
    title_cell.value = "Job Vacancy Reports" 

    title_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    title_cell.fill = PatternFill("solid", fgColor="1E1E1E") 
    title_cell.font = Font(color=colors.WHITE, bold=True, size=20)
    title_cell.alignment = Alignment(horizontal="center", vertical="center") 

     # must always start at 1
    for col in range(len(columns)):         
        ws.cell(row=2, column=col+1, value=columns[col])
        ws.row_dimensions[2].height = 30 
        ws.column_dimensions[alphabetic_chars[col]].width = 40

    header_columns = [ws.cell(row=2,column=x) for x in range(1,(len(columns)+1))]
    ft = Font(color=colors.BLACK, bold=True, size=12)
 
    for element in header_columns:
        element.font = ft  # the change only affects A1
        element.border = Border(top=thick, left=thick, right=thick, bottom=thick)
        element.fill = PatternFill("solid", fgColor="FFE300")
        element.alignment = Alignment(horizontal="center", vertical="center")

    # Database Query
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    job_vacancy = PRIJobVacancyInfo.objects.all().order_by('-id')

    index = 2
    for vacancy in job_vacancy:        
        index += 1        
        ws.cell(row=index, column=1, value=vacancy.id).border = thin_border
        ws.cell(row=index, column=2, value=vacancy.client_request.requested_client.client_company_name).border = thin_border
        ws.cell(row=index, column=3, value=vacancy.job_title).border = thin_border
        ws.cell(row=index, column=4, value=vacancy.job_category).border = thin_border
        ws.cell(row=index, column=5, value=vacancy.job_specialization).border = thin_border
        ws.cell(row=index, column=6, value=vacancy.job_minimum_experience).border = thin_border
        ws.cell(row=index, column=7, value=vacancy.job_salary).border = thin_border
        ws.cell(row=index, column=8, value=vacancy.job_deadline).border = thin_border


    workbook.save(response)

    return response

def export_report_xlsx_client_requests(request):

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename=client-requests-{date}.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

    workbook = Workbook()

    ws = workbook.active
    
    ws = workbook.create_sheet("Client Requests")

    ws.title = "Client Requests"

    ws.sheet_properties.tabColor = "1072BA"
    
    thin = Side(border_style="thin", color="007BFF")
    thick = Side(border_style="thick", color="1E1E1E")
    double = Side(border_style="double", color="ff0000") 

    # Adjust the size
    ws.row_dimensions[1].height = 30  
    
    columns = ["ID","Client","Content","Status","Date Requested"]
    alphabetic_chars = string.ascii_uppercase 
        
    # Title
    merge_cell = 'A1:{}1'.format(alphabetic_chars[len(columns)-1]) 

    ws.merge_cells(merge_cell) 
    title_cell = ws['A1'] 
    title_cell.value = "Client Request Reports" 

    title_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    title_cell.fill = PatternFill("solid", fgColor="1E1E1E") 
    title_cell.font = Font(color=colors.WHITE, bold=True, size=20)
    title_cell.alignment = Alignment(horizontal="center", vertical="center") 

     # must always start at 1
    for col in range(len(columns)):         
        ws.cell(row=2, column=col+1, value=columns[col])
        ws.row_dimensions[2].height = 30 
        ws.column_dimensions[alphabetic_chars[col]].width = 40

    header_columns = [ws.cell(row=2,column=x) for x in range(1,(len(columns)+1))]
    ft = Font(color=colors.BLACK, bold=True, size=12)
 
    for element in header_columns:
        element.font = ft  # the change only affects A1
        element.border = Border(top=thick, left=thick, right=thick, bottom=thick)
        element.fill = PatternFill("solid", fgColor="FFE300")
        element.alignment = Alignment(horizontal="center", vertical="center")

    # Database Query
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    job_requests = PRIRequestInfo.objects.all().order_by('-id')

    index = 2
    for job_request in job_requests:        
        index += 1    
        ws.cell(row=index, column=1, value=job_request.id).border = thin_border
        ws.cell(row=index, column=2, value=job_request.requested_client.client_company_name).border = thin_border
        ws.cell(row=index, column=3, value=job_request.content).border = thin_border
        ws.cell(row=index, column=4, value=job_request.status).border = thin_border
        ws.cell(row=index, column=5, value=job_request.data_requested).border = thin_border

    workbook.save(response)

    return response

def export_report_xlsx_hired_applicants(request):

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename=hired-applicants-{date}.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

    workbook = Workbook()

    ws = workbook.active
    
    ws = workbook.create_sheet("Hired Applicants")

    ws.title = "Hired Applicants"

    ws.sheet_properties.tabColor = "1072BA"

    thin = Side(border_style="thin", color="007BFF")
    thick = Side(border_style="thick", color="1E1E1E")
    double = Side(border_style="double", color="ff0000") 

    # Adjust the size
    ws.row_dimensions[1].height = 30  
    
    columns = ["ID","Company","Client Name","Location","Contact Person","Mobile","Email","ID","Position","Name","Mobile#","Email","Status"]
    alphabetic_chars = string.ascii_uppercase 
        
    # Title
    merge_cell = 'A1:{}1'.format(alphabetic_chars[len(columns)-1]) 

    ws.merge_cells(merge_cell) 
    title_cell = ws['A1'] 
    title_cell.value = "Hired Applicant Reports" 

    title_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    title_cell.fill = PatternFill("solid", fgColor="1E1E1E") 
    title_cell.font = Font(color=colors.WHITE, bold=True, size=20)
    title_cell.alignment = Alignment(horizontal="center", vertical="center") 

     # must always start at 1
    for col in range(len(columns)):         
        ws.cell(row=2, column=col+1, value=columns[col])
        ws.row_dimensions[2].height = 30 
        ws.column_dimensions[alphabetic_chars[col]].width = 40

    header_columns = [ws.cell(row=2,column=x) for x in range(1,(len(columns)+1))]
    ft = Font(color=colors.BLACK, bold=True, size=12)
 
    for element in header_columns:
        element.font = ft  # the change only affects A1
        element.border = Border(top=thick, left=thick, right=thick, bottom=thick)
        element.fill = PatternFill("solid", fgColor="FFE300")
        element.alignment = Alignment(horizontal="center", vertical="center")

    # Database Query
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    clients = PRIClientsInfo.objects.all().order_by('-id')

    index = 2
    for client in clients:        
        index += 1    
        ws.cell(row=index, column=1, value=client.id).border = thin_border
        ws.cell(row=index, column=2, value=client.client_company_name).border = thin_border
        ws.cell(row=index, column=3, value=str(client.client_user.last_name + ", " + client.client_user.first_name + " " + client.client_user.middle_name)).border = thin_border
        ws.cell(row=index, column=4, value=client.client_location).border = thin_border
        ws.cell(row=index, column=5, value=client.contact_person).border = thin_border
        ws.cell(row=index, column=6, value=client.contact_mobile_no).border = thin_border
        ws.cell(row=index, column=7, value=client.client_user.user.email).border = thin_border
        applicants = PRIApplicantJobHiredInfo.objects.all().filter(company_name=client).order_by('-id').distinct()
        for applicant in applicants:
            index += 1
            ws.cell(row=index, column=8, value=applicant.id).border = thin_border
            ws.cell(row=index, column=9, value=str(applicant.hired_applicant.lname + ", " + applicant.hired_applicant.fname + " " +applicant.hired_applicant.mname)).border = thin_border
            ws.cell(row=index, column=10, value=applicant.hired_applicant.mobile_no).border = thin_border
            ws.cell(row=index, column=11, value=applicant.hired_applicant.user.email).border = thin_border
            ws.cell(row=index, column=12, value=applicant.status).border = thin_border
            

    
    workbook.save(response)

    return response

def export_report_xlsx_clients(request):

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename=clients-{date}.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

    workbook = Workbook()

    ws = workbook.active
    
    ws = workbook.create_sheet("Clients")

    ws.title = "Clients"

    ws.sheet_properties.tabColor = "1072BA"

    thin = Side(border_style="thin", color="007BFF")
    thick = Side(border_style="thick", color="1E1E1E")
    double = Side(border_style="double", color="ff0000") 

    # Adjust the size
    ws.row_dimensions[1].height = 30  
    
    columns = ["ID","Company","Client Name","Location","Contact Person","Mobile","Email"]
    alphabetic_chars = string.ascii_uppercase 
        
    # Title
    merge_cell = 'A1:{}1'.format(alphabetic_chars[len(columns)-1]) 

    ws.merge_cells(merge_cell) 
    title_cell = ws['A1'] 
    title_cell.value = "Client Reports" 

    title_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    title_cell.fill = PatternFill("solid", fgColor="1E1E1E") 
    title_cell.font = Font(color=colors.WHITE, bold=True, size=20)
    title_cell.alignment = Alignment(horizontal="center", vertical="center") 

     # must always start at 1
    for col in range(len(columns)):         
        ws.cell(row=2, column=col+1, value=columns[col])
        ws.row_dimensions[2].height = 30 
        ws.column_dimensions[alphabetic_chars[col]].width = 40

    header_columns = [ws.cell(row=2,column=x) for x in range(1,(len(columns)+1))]
    ft = Font(color=colors.BLACK, bold=True, size=12)
 
    for element in header_columns:
        element.font = ft  # the change only affects A1
        element.border = Border(top=thick, left=thick, right=thick, bottom=thick)
        element.fill = PatternFill("solid", fgColor="FFE300")
        element.alignment = Alignment(horizontal="center", vertical="center")

    # Database Query
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    clients = PRIClientsInfo.objects.all().order_by('-id')

    index = 2
    for client in clients:        
        index += 1    
        ws.cell(row=index, column=1, value=client.id).border = thin_border
        ws.cell(row=index, column=2, value=client.client_company_name).border = thin_border
        ws.cell(row=index, column=3, value=str(client.client_user.last_name + ", " + client.client_user.first_name + " " + client.client_user.middle_name)).border = thin_border
        ws.cell(row=index, column=4, value=client.client_location).border = thin_border
        ws.cell(row=index, column=5, value=client.contact_person).border = thin_border
        ws.cell(row=index, column=6, value=client.contact_mobile_no).border = thin_border
        ws.cell(row=index, column=7, value=client.client_user.user.email).border = thin_border
  
    workbook.save(response)

    return response

def export_report_xlsx_current_users(request):

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = 'attachment; filename=current-users-{date}.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d'),)

    workbook = Workbook()

    ws = workbook.active
    
    ws = workbook.create_sheet("Current Users")

    ws.title = "Current Users"

    ws.sheet_properties.tabColor = "1072BA"

    thin = Side(border_style="thin", color="007BFF")
    thick = Side(border_style="thick", color="1E1E1E")
    double = Side(border_style="double", color="ff0000") 

    # Adjust the size
    ws.row_dimensions[1].height = 30  
    
    columns = ["ID", "Name", "Email", "Contact", "Position", "Department", "Level"]
    alphabetic_chars = string.ascii_uppercase 
        
    # Title
    merge_cell = 'A1:{}1'.format(alphabetic_chars[len(columns)-1]) 

    ws.merge_cells(merge_cell) 
    title_cell = ws['A1'] 
    title_cell.value = "User Reports" 

    title_cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)
    title_cell.fill = PatternFill("solid", fgColor="1E1E1E") 
    title_cell.font = Font(color=colors.WHITE, bold=True, size=20)
    title_cell.alignment = Alignment(horizontal="center", vertical="center") 

     # must always start at 1
    for col in range(len(columns)):         
        ws.cell(row=2, column=col+1, value=columns[col])
        ws.row_dimensions[2].height = 30 
        ws.column_dimensions[alphabetic_chars[col]].width = 40

    header_columns = [ws.cell(row=2,column=x) for x in range(1,(len(columns)+1))]
    ft = Font(color=colors.BLACK, bold=True, size=12)
 
    for element in header_columns:
        element.font = ft  # the change only affects A1
        element.border = Border(top=thick, left=thick, right=thick, bottom=thick)
        element.fill = PatternFill("solid", fgColor="FFE300")
        element.alignment = Alignment(horizontal="center", vertical="center")

    # Database Query
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

    users_info = PRIUserInfo.objects.all().order_by('-id')

    index = 2
    for users in users_info:        
        index += 1    
        ws.cell(row=index, column=1, value=users.id).border = thin_border
        ws.cell(row=index, column=2, value=str(users.last_name + ", " + users.first_name + " " + users.middle_name)).border = thin_border
        ws.cell(row=index, column=3, value=users.user.email).border = thin_border
        ws.cell(row=index, column=4, value=users.contact).border = thin_border
        ws.cell(row=index, column=5, value=users.position).border = thin_border
        ws.cell(row=index, column=6, value=users.department).border = thin_border
        ws.cell(row=index, column=7, value=users.level).border = thin_border
    

    workbook.save(response)

    return response